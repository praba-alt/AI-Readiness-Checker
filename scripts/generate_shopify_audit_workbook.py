#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import math
from copy import copy
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


SITE_LABELS = {
    "trailberg.com": "Trailberg",
    "lorenzo.world": "Lorenzo",
    "dreamisfree.com": "Dream Is Free",
    "aariclothing.com": "Aari Clothing",
}


SUMMARY_ROWS = [
    (
        "Stage 0 - Visibility",
        "AI Brand Visibility",
        "Public AI brand visibility tracking",
        "Integrated brand visibility report",
        "Covered by the in-repo brand visibility report, then merged into the master workbook.",
        "out_of_scope",
    ),
    (
        "Stage 1 - Foundation",
        "Homepage Crawlability",
        "Homepage returns 200, is indexable, and exposes a clear title/H1/meta description.",
        "Homepage source code",
        "LLMs and search systems need a stable, machine-readable homepage to anchor brand understanding.",
        "homepage_foundation",
    ),
    (
        "Stage 1 - Foundation",
        "Robots + Sitemap",
        "robots.txt and XML sitemap are available and intentional.",
        "Technical SEO",
        "Crawlers need clean discovery paths before deeper AI-ready signals matter.",
        "crawl_discovery",
    ),
    (
        "Stage 1 - Foundation",
        "URL Structure",
        "Products, collections, pages, and blogs are separated cleanly in public URLs.",
        "Information architecture",
        "Clear entities reduce ambiguity for search engines and retrieval systems.",
        "url_structure",
    ),
    (
        "Stage 1 - Foundation",
        "Collection Discovery",
        "Collection/category pages are discoverable and have their own metadata.",
        "Collection pages",
        "Collections teach category semantics and shopping context to AI systems.",
        "collection_foundation",
    ),
    (
        "Stage 1 - Foundation",
        "Product Metadata",
        "Product pages expose strong titles, canonical URLs, H1s, and descriptions.",
        "Product pages",
        "This determines whether PDPs can be cited and compared correctly.",
        "product_metadata",
    ),
    (
        "Stage 1 - Foundation",
        "Structured Product Data",
        "Product schema exposes machine-readable product facts such as offers and brand.",
        "JSON-LD / schema",
        "Structured product facts are the strongest public readiness signal for AI retrieval.",
        "product_schema",
    ),
    (
        "Stage 1 - Foundation",
        "Machine-readable Feeds",
        "The storefront exposes products.json or equivalent public feeds.",
        "Shopify feeds",
        "Feeds help external systems consume products without scraping every PDP.",
        "machine_feeds",
    ),
    (
        "Stage 2 - Agentic Value",
        "Recommendation Context",
        "Sample PDPs expose enough visible context for recommendations, FAQs, or sizing/support cues.",
        "Product copy + support content",
        "Agents perform better when storefront copy answers follow-up questions directly.",
        "agent_context",
    ),
    (
        "Stage 2 - Agentic Value",
        "Breadcrumbs + Linking",
        "Sample pages expose breadcrumb schema or clear internal navigation paths.",
        "Navigation semantics",
        "This improves entity relationships between brand, category, and product.",
        "breadcrumbs",
    ),
    (
        "Stage 2 - Agentic Value",
        "Search + Facets",
        "Search/facet behaviour is visible from the public storefront.",
        "Search UX",
        "Facet and search signals expose how products can be filtered and reasoned over.",
        "search_and_facets",
    ),
    (
        "Stage 2 - Agentic Value",
        "Trust + Support Content",
        "About, policy, and help content is discoverable from the storefront.",
        "Brand/support content",
        "LLMs need accessible trust signals beyond product grids.",
        "trust_content",
    ),
    (
        "Stage 3 - Ecosystem",
        "llms.txt",
        "The site publishes llms.txt or a well-known equivalent.",
        "AI access policy",
        "This is an emerging but explicit AI-readiness signal.",
        "llms_txt",
    ),
    (
        "Stage 3 - Ecosystem",
        "AI Bot Policy",
        "robots.txt does not unintentionally block major AI crawlers.",
        "robots.txt",
        "AI-facing discovery depends on whether bots can access content at all.",
        "ai_bot_policy",
    ),
    (
        "Stage 3 - Ecosystem",
        "Native Shopify MCP",
        "Shopify Storefront MCP is live and reachable.",
        "Shopify MCP / tool access",
        "Shopify's built-in MCP gives agents a base commerce layer for catalog, cart, and policy flows.",
        "mcp_discovery",
    ),
    (
        "Stage 3 - Ecosystem",
        "Regional Signals",
        "The site exposes market or locale signals such as hreflang or multiple currencies.",
        "International SEO",
        "Regional context helps AI systems quote the correct pricing and availability context.",
        "regional_signals",
    ),
]


DETAIL_CHECKS = [
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Crawlability",
        "question": "Homepage returns 200 and is indexable.",
        "why": "This is the root entity page for both search crawlers and LLM retrieval.",
        "key": "homepage_status",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Content",
        "domain": "Crawlability",
        "question": "Homepage title, meta description, and H1 are all present and reasonably sized.",
        "why": "Weak homepage semantics often propagate into poor AI brand framing.",
        "key": "homepage_metadata",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Structured Data",
        "question": "Homepage exposes Organization or WebSite schema.",
        "why": "This gives machines an explicit brand and site entity to anchor on.",
        "key": "homepage_schema",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Structured Data",
        "question": "Homepage schema is deep enough for intent-driven discovery (brand entity, site entity, search action, sameAs/identity cues).",
        "why": "Presence alone is weaker than rich entity semantics for retrieval and intent matching.",
        "key": "homepage_schema_depth",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Crawlability",
        "question": "robots.txt exists and points crawlers at the sitemap.",
        "why": "Crawler instructions should be explicit, not left to inference.",
        "key": "robots",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Crawlability",
        "question": "XML sitemaps enumerate products and collections.",
        "why": "This exposes clean entity lists for products and category pages.",
        "key": "sitemap",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "SEO",
        "domain": "URL Structure",
        "question": "Public URLs separate products, collections, pages, and blogs cleanly.",
        "why": "Consistent URL entities reduce ambiguity for external systems.",
        "key": "url_structure",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "SEO",
        "domain": "Collections",
        "question": "Collection/category pages are publicly discoverable.",
        "why": "Collections carry category semantics that PDPs do not.",
        "key": "collections_exist",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "SEO",
        "domain": "Collections",
        "question": "Sample collection pages have title, meta description, H1, and canonical coverage.",
        "why": "Collection metadata often influences category understanding in AI-generated answers.",
        "key": "collection_metadata",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Content",
        "domain": "Collections",
        "question": "Sample collection pages include enough visible copy or navigation context beyond product tiles.",
        "why": "Thin collection pages weaken category comprehension and comparison quality.",
        "key": "collection_content",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "SEO",
        "domain": "Products",
        "question": "Sample product pages have title, meta description, H1, and canonical coverage.",
        "why": "PDP metadata quality directly affects citation and recommendation eligibility.",
        "key": "product_metadata",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Products",
        "question": "Sample product pages expose Product JSON-LD.",
        "why": "Product schema is the clearest public signal of machine-readable commerce data.",
        "key": "product_schema_presence",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Products",
        "question": "Product schema includes offers, brand, and product description fields.",
        "why": "Schema without commercial fields is less useful for retrieval and comparison.",
        "key": "product_schema_completeness",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Products",
        "question": "Product and collection schemas are deep enough for intent-driven discovery, not just shallow presence.",
        "why": "Intent-driven AI ranking benefits from richer fields such as category, material, audience, ItemList, and commercial attributes.",
        "key": "schema_depth",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "SEO",
        "domain": "Products",
        "question": "Sample product pages expose breadcrumb schema or equivalent hierarchy signals.",
        "why": "Breadcrumbs tie products back to collection-level meaning.",
        "key": "breadcrumbs",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Content",
        "domain": "Media Semantics",
        "question": "Sample pages do not show widespread missing alt text on images.",
        "why": "Alt text still matters for accessibility and secondary machine interpretation.",
        "key": "image_alt",
    },
    {
        "phase": "Stage 1 - Foundation",
        "team": "Dev",
        "domain": "Feeds",
        "question": "products.json is publicly available.",
        "why": "Public feeds reduce scraping friction for structured product ingestion.",
        "key": "products_json",
    },
    {
        "phase": "Stage 2 - Agentic Value",
        "team": "Content",
        "domain": "Product Semantics",
        "question": "Sample PDPs show semantic depth such as FAQ, reviews, or sizing/support cues.",
        "why": "Agents need follow-up-answer material, not only marketing headlines.",
        "key": "semantic_depth",
    },
    {
        "phase": "Stage 2 - Agentic Value",
        "team": "Dev",
        "domain": "Search",
        "question": "The storefront exposes search entry points and collection-page facet signals.",
        "why": "Facet/search visibility hints at how well the catalogue can be filtered or reasoned over.",
        "key": "search_facets",
    },
    {
        "phase": "Stage 2 - Agentic Value",
        "team": "Content",
        "domain": "Trust Content",
        "question": "About, policy, or help content is discoverable from the public site.",
        "why": "Agents need trust and support context alongside commercial pages.",
        "key": "trust_content",
    },
    {
        "phase": "Stage 2 - Agentic Value",
        "team": "Data",
        "domain": "Conversational Readiness",
        "question": "The public storefront reveals enough structured facts to answer basic follow-up questions.",
        "why": "If public pages are thin, AI systems must guess or ignore the site.",
        "key": "follow_up_readiness",
    },
    {
        "phase": "Stage 2 - Agentic Value",
        "team": "Merchandising",
        "domain": "Product Relationships",
        "question": "Sample PDPs visibly expose related products, alternatives, or cross-sell modules.",
        "why": "Intent-driven recommendations improve when the storefront teaches product relationships explicitly.",
        "key": "related_products",
    },
    {
        "phase": "Stage 2 - Agentic Value",
        "team": "Content",
        "domain": "Clothing Attributes",
        "question": "Sample clothing PDPs expose deep field signals such as fabric, fit, weather, style, care, and use-case context.",
        "why": "For apparel, these are the high-intent fields agents need to match products to real user intent.",
        "key": "clothing_attribute_depth",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "Dev",
        "domain": "AI Access Policy",
        "question": "llms.txt is published.",
        "why": "This is an explicit but optional AI-discovery signal.",
        "key": "llms_txt",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "Dev",
        "domain": "AI Access Policy",
        "question": "robots.txt does not block major AI crawlers.",
        "why": "Blocking AI crawlers undermines discoverability regardless of other SEO quality.",
        "key": "ai_bot_policy",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "Data",
        "domain": "Feeds",
        "question": "The site exposes machine-readable data beyond HTML (schema + feed).",
        "why": "Multiple structured channels make ingestion more reliable.",
        "key": "machine_feeds",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "Dev",
        "domain": "Structured Data",
        "question": "Visible page content and schema appear broadly aligned on sampled PDPs.",
        "why": "Misaligned human-visible and machine-readable content erodes trust in the data.",
        "key": "schema_alignment",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "SEO",
        "domain": "Regionalisation",
        "question": "Locale or market signals such as hreflang/currency are publicly visible.",
        "why": "Regional context is important for correct AI answers around price and availability.",
        "key": "regional_signals",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "Dev",
        "domain": "MCP",
        "question": "Shopify Storefront MCP is live and discoverable.",
        "why": "This gives agents a native commerce layer for catalog search, cart operations, and policy answers.",
        "key": "mcp",
    },
    {
        "phase": "Stage 3 - Ecosystem",
        "team": "Dev",
        "domain": "MCP",
        "question": "Shopify Storefront MCP returns a usable tool list and RPC response.",
        "why": "A working native MCP should expose stable commerce tools, not just an endpoint that exists.",
        "key": "mcp_depth",
    },
]


def result(status: str, score: int, evidence: str, recommendation: str, observability: str = "Public storefront", sources: str = "") -> Dict[str, Any]:
    return {
        "status": status,
        "score": score,
        "evidence": evidence,
        "recommendation": recommendation,
        "observability": observability,
        "sources": sources,
    }


def avg(values: List[float]) -> float:
    return round(mean(values), 2) if values else 0.0


def site_label(site: Dict[str, Any]) -> str:
    return SITE_LABELS.get(site["domain"], site["domain"])


def build_metrics(site: Dict[str, Any]) -> Dict[str, Any]:
    homepage = site["homepage"]
    robots = site["robots"]
    sitemap = site["sitemap"]
    derived = site["derived"]
    ai = site["ai_readiness"]
    products = site["samples"]["product_pages"]
    collections = site["samples"]["collection_pages"]
    content = site["samples"]["content_pages"]

    product_fields = []
    for page in products:
        product_fields.extend(page.get("product_schema_fields", []))

    home_links = homepage.get("sample_internal_links", [])
    about_link = any("/about" in link.lower() or "/pages/about" in link.lower() for link in home_links)
    policy_link = any(
        marker in link.lower()
        for link in home_links
        for marker in ("/policies/", "/privacy", "/refund", "/shipping", "/terms", "/help", "/support")
    ) or homepage.get("mentions_policy", False)

    collection_summary = derived["collection_summary"]
    product_summary = derived["product_summary"]
    content_summary = derived["content_summary"]
    mcp_manifest = ai.get("mcp_manifest", [])
    mcp_rpc = ai.get("mcp_rpc", [])
    clothing_categories = product_summary.get("clothing_attribute_categories", [])
    mcp_tool_count = max((item.get("tool_count", 0) for item in mcp_manifest), default=0)
    mcp_endpoint = next(
        (
            item.get("final_url", "")
            for item in mcp_manifest
            if item.get("tool_count", 0) > 0 and item.get("final_url")
        ),
        next(
            (
                item.get("url", "")
                for item in mcp_rpc
                if item.get("tool_count", 0) > 0 and item.get("url")
            ),
            "",
        ),
    )
    mcp_protocol_version = next(
        (
            item.get("protocol_version", "")
            for item in mcp_rpc
            if item.get("tool_count", 0) > 0 and item.get("protocol_version")
        ),
        "",
    )
    support_urls = derived.get("support_urls", [])
    rendered_alt_missing = product_summary.get("avg_product_gallery_images_without_alt", 0)
    rendered_alt_images = product_summary.get("avg_product_gallery_image_count", 0)
    support_signal_total = (
        product_summary.get("support_link_total", 0)
        + content_summary.get("support_link_total", 0)
        + content_summary.get("policy_link_total", 0)
    )
    follow_up_signal_count = (
        product_summary.get("faq_heading_total", 0)
        + product_summary.get("size_guide_link_total", 0)
        + product_summary.get("review_signal_total", 0)
        + product_summary.get("related_module_count", 0)
    )

    return {
        "homepage_ok": bool(homepage.get("ok")),
        "homepage_indexable": "noindex" not in (homepage.get("meta_robots") or "").lower(),
        "homepage_title_ok": 20 <= homepage.get("title_length", 0) <= 70,
        "homepage_meta_ok": 50 <= homepage.get("meta_description_length", 0) <= 180,
        "homepage_single_h1": homepage.get("h1_count", 0) == 1,
        "homepage_schema": homepage.get("organization_schema_count", 0) > 0 or homepage.get("website_schema_count", 0) > 0,
        "homepage_schema_depth_ok": (
            homepage.get("organization_schema_count", 0) > 0
            and homepage.get("website_schema_count", 0) > 0
            and (
                homepage.get("has_search_action_schema", False)
                or homepage.get("search_action_schema_count", 0) > 0
            )
        ),
        "robots_ok": bool(robots.get("ok")),
        "robots_has_sitemap": bool((robots.get("parsed") or {}).get("sitemap_directives")),
        "ai_bot_blocks": ai.get("ai_bot_blocks", []),
        "sitemap_ok": sitemap.get("total_urls", 0) > 0,
        "product_count": sitemap["counts"].get("product", 0),
        "collection_count": sitemap["counts"].get("collection", 0),
        "url_query_ratio": sitemap["url_patterns"].get("query_ratio", 0),
        "collections_sampled": collection_summary.get("sampled", 0),
        "collections_ok": collection_summary.get("ok", 0),
        "collection_meta_ok": (
            collection_summary.get("ok", 0) > 0
            and collection_summary.get("avg_meta_description_length", 0) >= 40
            and collection_summary.get("all_have_h1", False)
            and collection_summary.get("canonical_match_ratio", 0) >= 0.66
        ),
        "collection_content_ok": (
            collection_summary.get("avg_word_count", 0) >= 80
            or collection_summary.get("avg_collection_intro_word_count", 0) >= 80
        ),
        "products_sampled": product_summary.get("sampled", 0),
        "products_ok": product_summary.get("ok", 0),
        "product_meta_ok": (
            product_summary.get("ok", 0) > 0
            and product_summary.get("avg_meta_description_length", 0) >= 40
            and product_summary.get("all_have_h1", False)
            and product_summary.get("canonical_match_ratio", 0) >= 0.66
        ),
        "product_schema_all": product_summary.get("all_have_product_schema", False),
        "avg_product_schema_depth": product_summary.get("avg_product_schema_depth", 0),
        "collection_schema_depth_ok": collection_summary.get("all_have_collection_semantics", False),
        "breadcrumb_all": product_summary.get("all_have_breadcrumb_schema", False),
        "faq_any": (
            product_summary.get("any_have_faq_schema", False)
            or any(page.get("mentions_faq") for page in products)
            or product_summary.get("faq_heading_total", 0) > 0
        ),
        "size_guide_any": (
            any(page.get("mentions_size_guide") for page in products)
            or product_summary.get("size_guide_link_total", 0) > 0
        ),
        "reviews_any": (
            any(page.get("mentions_reviews") for page in products)
            or product_summary.get("review_signal_total", 0) > 0
        ),
        "product_schema_offer_ratio": avg([1 if item.get("has_offers") else 0 for item in product_fields]),
        "product_schema_brand_ratio": avg([1 if item.get("has_brand") else 0 for item in product_fields]),
        "product_schema_desc_ratio": avg([1 if item.get("has_description") else 0 for item in product_fields]),
        "avg_missing_alt": avg([page.get("images_without_alt", 0) for page in products + collections]),
        "avg_rendered_gallery_missing_alt": rendered_alt_missing,
        "avg_rendered_gallery_image_count": rendered_alt_images,
        "avg_rendered_gallery_alt_match_ratio": product_summary.get("avg_product_gallery_alt_match_ratio", 0),
        "products_json_ok": bool(ai["products_json"].get("ok")),
        "products_json_titles": ai["products_json"].get("sample_product_titles", []),
        "about_link": about_link,
        "policy_link": policy_link or content_summary.get("policy_link_total", 0) > 0,
        "support_urls": support_urls,
        "support_signal_total": support_signal_total,
        "content_pages_sampled": content_summary.get("sampled", 0),
        "has_search_link": homepage.get("has_search_link", False) or homepage.get("search_form_count", 0) > 0,
        "predictive_search_signal": homepage.get("predictive_search_markers", 0) > 0,
        "facet_signal": (
            any(page.get("mentions_filter") or page.get("mentions_sort") for page in collections)
            or collection_summary.get("filter_control_total", 0) > 0
            or collection_summary.get("sort_control_total", 0) > 0
        ),
        "follow_up_signal": (
            any(page.get("word_count", 0) >= 180 for page in products)
            and follow_up_signal_count > 0
        ),
        "follow_up_signal_count": follow_up_signal_count,
        "related_products_any": (
            any(page.get("has_related_products_module") for page in products)
            or product_summary.get("related_module_count", 0) > 0
        ),
        "clothing_attribute_categories": clothing_categories,
        "clothing_attribute_depth_ok": len(clothing_categories) >= 3,
        "llms_ok": bool(derived.get("has_any_llms")),
        "mcp_ok": bool(derived.get("has_public_mcp")),
        "mcp_rpc_ok": bool(derived.get("has_working_mcp_rpc")),
        "mcp_tool_count": mcp_tool_count,
        "mcp_endpoint": mcp_endpoint,
        "mcp_protocol_version": mcp_protocol_version,
        "machine_feeds_ok": bool(ai["products_json"].get("ok")) and product_summary.get("all_have_product_schema", False),
        "regional_signal": (
            homepage.get("has_hreflang", False)
            or bool(derived.get("currencies"))
            or bool(derived.get("country_codes"))
            or bool(product_summary.get("currency_signals"))
            or bool(content_summary.get("currency_signals"))
            or product_summary.get("any_region_selector", False)
            or content_summary.get("any_region_selector", False)
        ),
        "schema_alignment_ok": (
            product_summary.get("all_have_product_schema", False)
            and product_summary.get("canonical_match_ratio", 0) >= 0.66
            and product_summary.get("meta_description_matches_body_prefix_count", 0) >= max(1, product_summary.get("ok", 0) // 2)
        ),
        "home_links": home_links,
        "homepage": homepage,
        "robots": robots,
        "sitemap": sitemap,
        "products": products,
        "collections": collections,
        "content": content,
        "derived": derived,
        "mcp_manifest": mcp_manifest,
        "mcp_rpc": mcp_rpc,
        "site": site,
    }


def eval_check(key: str, metrics: Dict[str, Any]) -> Dict[str, Any]:
    homepage = metrics["homepage"]
    sitemap = metrics["sitemap"]
    if key == "homepage_status":
        if metrics["homepage_ok"] and metrics["homepage_indexable"]:
            return result("PASS", 5, f"Homepage {homepage.get('final_url')} returned {homepage.get('status')} and no noindex was detected.", "Keep homepage crawlable and monitor unexpected noindex directives.", sources="homepage")
        if metrics["homepage_ok"]:
            return result("PARTIAL", 3, "Homepage responded, but noindex or conflicting robots directives were detected.", "Review homepage indexation directives.", sources="homepage")
        return result("FAIL", 0, f"Homepage fetch failed with status {homepage.get('status')} / error {homepage.get('fetch_error')}.", "Fix homepage accessibility before deeper optimisation.", sources="homepage")

    if key == "homepage_metadata":
        checks = [
            metrics["homepage_title_ok"],
            metrics["homepage_meta_ok"],
            metrics["homepage_single_h1"],
        ]
        passed = sum(1 for item in checks if item)
        evidence = f"title_len={homepage.get('title_length')}, meta_len={homepage.get('meta_description_length')}, h1_count={homepage.get('h1_count')}"
        if passed == 3:
            return result("PASS", 5, evidence, "Maintain strong homepage metadata hygiene.", sources="homepage")
        if passed >= 1:
            return result("PARTIAL", 3, evidence, "Tighten homepage title/meta/H1 consistency.", sources="homepage")
        return result("FAIL", 1, evidence, "Rewrite homepage metadata and heading structure.", sources="homepage")

    if key == "homepage_schema":
        if metrics["homepage_schema"]:
            return result("PASS", 5, f"Homepage schema types: {', '.join(homepage.get('json_ld_types', [])[:6]) or 'none listed'}.", "Keep Organization/WebSite schema aligned with visible brand facts.", sources="homepage")
        return result("FAIL", 1, "No Organization/WebSite schema was detected on the homepage.", "Add Organization and WebSite JSON-LD to the homepage.", sources="homepage")

    if key == "homepage_schema_depth":
        evidence = (
            f"org={homepage.get('organization_schema_count', 0)}, "
            f"website={homepage.get('website_schema_count', 0)}, "
            f"search_action={homepage.get('search_action_schema_count', 0) or int(homepage.get('has_search_action_schema', False))}"
        )
        if metrics["homepage_schema_depth_ok"]:
            return result("PASS", 5, evidence, "Maintain rich homepage entity schema with search-action and identity cues.", sources="homepage schema")
        if metrics["homepage_schema"]:
            return result("PARTIAL", 3, evidence, "Deepen homepage schema beyond basic presence with search action and stronger identity fields.", sources="homepage schema")
        return result("FAIL", 1, evidence, "Homepage schema is too shallow or absent for intent-driven discovery.", sources="homepage schema")

    if key == "robots":
        directives = (metrics["robots"].get("parsed") or {}).get("sitemap_directives", [])
        if metrics["robots_ok"] and metrics["robots_has_sitemap"]:
            return result("PASS", 5, f"robots.txt is present and declares {len(directives)} sitemap directive(s).", "Keep robots rules intentional and reviewed after theme/app changes.", sources="robots.txt")
        if metrics["robots_ok"]:
            return result("PARTIAL", 3, "robots.txt is present but sitemap directives were not found.", "Declare sitemap.xml explicitly in robots.txt.", sources="robots.txt")
        return result("FAIL", 0, "robots.txt could not be fetched.", "Publish a valid robots.txt file.", sources="robots.txt")

    if key == "sitemap":
        if metrics["sitemap_ok"] and metrics["product_count"] > 0 and metrics["collection_count"] > 0:
            return result("PASS", 5, f"sitemap.xml exposed {metrics['product_count']} product URLs and {metrics['collection_count']} collection URLs.", "Keep product and collection sitemap coverage intact.", sources="sitemap.xml")
        if metrics["sitemap_ok"]:
            return result("PARTIAL", 3, f"sitemap.xml exists but coverage looked thin: products={metrics['product_count']}, collections={metrics['collection_count']}.", "Check sitemap coverage for missing product or collection templates.", sources="sitemap.xml")
        return result("FAIL", 0, "No usable XML sitemap was discovered.", "Publish a valid XML sitemap.", sources="sitemap.xml")

    if key == "url_structure":
        evidence = f"products={metrics['product_count']}, collections={metrics['collection_count']}, query_ratio={metrics['url_query_ratio']}"
        if metrics["product_count"] and metrics["collection_count"] and metrics["url_query_ratio"] <= 0.05:
            return result("PASS", 5, evidence, "Preserve clean entity-specific URL patterns.", sources="sitemap.xml")
        if metrics["product_count"] or metrics["collection_count"]:
            return result("PARTIAL", 3, evidence, "Reduce noisy parameter URLs and ensure all key entity types are exposed.", sources="sitemap.xml")
        return result("FAIL", 1, evidence, "The URL inventory is too thin to prove clean storefront entity separation.", sources="sitemap.xml")

    if key == "collections_exist":
        if metrics["collections_ok"] >= 1:
            return result("PASS", 5, f"Sampled {metrics['collections_ok']} collection pages successfully.", "Retain clean public collection landing pages.", sources="collection samples")
        return result("FAIL", 1, "No collection pages were successfully sampled from the sitemap.", "Ensure collection/category pages are discoverable and crawlable.", sources="collection samples")

    if key == "collection_metadata":
        if metrics["collection_meta_ok"]:
            return result("PASS", 5, "Sample collection pages had H1, metadata, and canonical coverage.", "Keep collection metadata consistent across the catalogue.", sources="collection samples")
        if metrics["collections_ok"] >= 1:
            return result("PARTIAL", 3, "Collection pages exist, but metadata/canonical coverage was inconsistent across samples.", "Improve collection title/meta/H1/canonical consistency.", sources="collection samples")
        return result("FAIL", 1, "Collection-page metadata could not be validated from successful samples.", "Audit collection templates for missing title/meta/H1/canonical tags.", sources="collection samples")

    if key == "collection_content":
        if metrics["collection_content_ok"]:
            return result("PASS", 4, f"Average sampled collection word count was {metrics['derived']['collection_summary'].get('avg_word_count')} words.", "Keep category copy and navigation context visible.", sources="collection samples")
        if metrics["collections_ok"] >= 1:
            return result("PARTIAL", 2, f"Average sampled collection word count was {metrics['derived']['collection_summary'].get('avg_word_count')} words.", "Add stronger collection intro copy and semantic navigation context.", sources="collection samples")
        return result("FAIL", 1, "No collection content could be sampled reliably.", "Expose crawlable collection pages with supporting copy.", sources="collection samples")

    if key == "product_metadata":
        if metrics["product_meta_ok"]:
            return result("PASS", 5, "Sample product pages had H1, metadata, and canonical coverage.", "Maintain PDP metadata consistency.", sources="product samples")
        if metrics["products_ok"] >= 1:
            return result("PARTIAL", 3, "Product pages were reachable, but metadata/canonical signals were inconsistent.", "Tighten PDP title/meta/H1/canonical coverage.", sources="product samples")
        return result("FAIL", 1, "Product-page metadata could not be validated from successful samples.", "Audit PDP templates for crawlable metadata.", sources="product samples")

    if key == "product_schema_presence":
        if metrics["product_schema_all"]:
            return result("PASS", 5, "All sampled PDPs exposed Product JSON-LD.", "Keep Product schema deployed across every PDP template.", sources="product samples")
        if metrics["products_ok"] >= 1:
            return result("PARTIAL", 3, "Only some sampled PDPs exposed Product JSON-LD.", "Ensure Product schema is present on every PDP.", sources="product samples")
        return result("FAIL", 1, "No sampled PDPs exposed Product JSON-LD.", "Add Product JSON-LD to PDP templates.", sources="product samples")

    if key == "product_schema_completeness":
        ratios = (
            metrics["product_schema_offer_ratio"],
            metrics["product_schema_brand_ratio"],
            metrics["product_schema_desc_ratio"],
        )
        evidence = "offers={:.2f}, brand={:.2f}, description={:.2f}".format(*ratios)
        if min(ratios) >= 0.8:
            return result("PASS", 5, evidence, "Keep schema fields aligned with PDP content.", sources="product schema")
        if max(ratios) > 0:
            return result("PARTIAL", 3, evidence, "Expand Product schema to cover offers, brand, and descriptions consistently.", sources="product schema")
        return result("FAIL", 1, evidence, "Product schema was missing key commercial fields in sampled pages.", "Audit PDP schema output for missing fields.", sources="product schema")

    if key == "schema_depth":
        evidence = (
            f"product_depth={metrics['avg_product_schema_depth']}, "
            f"collection_semantics={metrics['collection_schema_depth_ok']}, "
            f"attribute_categories={len(metrics['clothing_attribute_categories'])}"
        )
        if metrics["avg_product_schema_depth"] >= 0.7 and metrics["collection_schema_depth_ok"]:
            return result("PASS", 5, evidence, "Keep schema depth aligned to commercial and intent-rich product attributes.", sources="schema depth")
        if metrics["product_schema_all"] or metrics["collection_schema_depth_ok"]:
            return result("PARTIAL", 3, evidence, "Deepen Product/Collection schema with richer attribute and list semantics.", sources="schema depth")
        return result("FAIL", 1, evidence, "Schemas appear present-but-shallow or inconsistent for intent-driven retrieval.", "Expand schema depth on product and collection templates.", sources="schema depth")

    if key == "breadcrumbs":
        if metrics["breadcrumb_all"]:
            return result("PASS", 4, "Breadcrumb schema was present across sampled PDPs.", "Keep hierarchy signals intact across templates.", sources="product samples")
        if metrics["products_ok"] >= 1:
            return result("PARTIAL", 2, "Breadcrumb schema was inconsistent across sampled PDPs.", "Add breadcrumb schema and visible breadcrumb trails.", sources="product samples")
        return result("FAIL", 1, "Breadcrumb coverage could not be validated.", "Add breadcrumb trails to product and collection pages.", sources="product samples")

    if key == "image_alt":
        rendered_missing = metrics["avg_rendered_gallery_missing_alt"]
        rendered_count = metrics["avg_rendered_gallery_image_count"]
        if rendered_count > 0 and rendered_missing <= 0.5:
            return result("PASS", 4, f"Sampled PDP gallery media averaged {rendered_count} images with {rendered_missing} missing alt fields; most rendered alts were present.", "Keep commerce-image alt text rendered consistently and improve image-specific descriptions over generic title fallbacks.", sources="rendered product samples")
        if rendered_count > 0:
            return result("PARTIAL", 2, f"Sampled PDP gallery media averaged {rendered_count} images with {rendered_missing} missing alt fields.", "Add alt-text QA to product and collection publishing workflows.", sources="rendered product samples")
        avg_missing = metrics["avg_missing_alt"]
        if avg_missing <= 1:
            return result("PASS", 4, f"Average missing-alt count across sampled product/collection pages was {avg_missing}.", "Continue enforcing alt text on commerce imagery.", sources="product+collection samples")
        if avg_missing <= 5:
            return result("PARTIAL", 2, f"Average missing-alt count across sampled product/collection pages was {avg_missing}.", "Audit template and CMS processes for missing alt text.", sources="product+collection samples")
        return result("FAIL", 1, f"Average missing-alt count across sampled product/collection pages was {avg_missing}.", "Add alt-text QA to product and collection publishing workflows.", sources="product+collection samples")

    if key == "products_json":
        if metrics["products_json_ok"]:
            return result("PASS", 5, f"products.json was reachable with sample titles: {', '.join(metrics['products_json_titles'][:3])}.", "Keep public product feeds intentional and rate-limited as needed.", sources="products.json")
        return result("FAIL", 1, "products.json was not publicly reachable.", "Decide whether to expose or proxy a public machine-readable feed.", sources="products.json")

    if key == "semantic_depth":
        signals = sum(
            [
                1 if metrics["faq_any"] else 0,
                1 if metrics["size_guide_any"] else 0,
                1 if metrics["reviews_any"] else 0,
            ]
        )
        evidence = (
            f"faq={metrics['faq_any']}, size_guide={metrics['size_guide_any']}, "
            f"reviews={metrics['reviews_any']}, follow_up_signals={metrics['follow_up_signal_count']}"
        )
        if signals >= 2:
            return result("PASS", 4, evidence, "Retain PDP support content that answers follow-up questions.", sources="product samples")
        if signals == 1:
            return result("PARTIAL", 2, evidence, "Add richer FAQ, sizing, or review/support cues on PDPs.", sources="product samples")
        return result("FAIL", 1, evidence, "Sample PDPs looked thin on FAQ/support/review cues.", "Add public follow-up-answer content to PDP templates.", sources="product samples")

    if key == "search_facets":
        evidence = (
            f"search={metrics['has_search_link']}, predictive_search={metrics['predictive_search_signal']}, "
            f"facets={metrics['facet_signal']}"
        )
        if metrics["has_search_link"] and metrics["facet_signal"]:
            return result("PASS", 4, evidence, "Maintain visible search and facet UX.", sources="homepage+collection samples")
        if metrics["has_search_link"] or metrics["facet_signal"]:
            return result("PARTIAL", 2, evidence, "Expose clearer search entry points and filter semantics.", sources="homepage+collection samples")
        return result("FAIL", 1, "No obvious public search or facet signals were detected in sampled pages.", "Review search discoverability and collection filter rendering.", sources="homepage+collection samples")

    if key == "trust_content":
        evidence = (
            f"about_link={metrics['about_link']}, policy_link={metrics['policy_link']}, "
            f"sampled_content={metrics['content_pages_sampled']}, support_urls={len(metrics['support_urls'])}"
        )
        if metrics["about_link"] and metrics["policy_link"] and len(metrics["support_urls"]) >= 2:
            return result("PASS", 4, evidence, "Keep trust and support content prominently linked.", sources="homepage+content samples")
        if metrics["about_link"] or metrics["policy_link"] or metrics["content_pages_sampled"] > 0:
            return result("PARTIAL", 2, evidence, "Improve discoverability of about/help/policy content.", sources="homepage+content samples")
        return result("FAIL", 1, "Trust/support content was not easily discoverable from sampled public paths.", "Link about/help/policy content more clearly in navigation or footer.", sources="homepage+content samples")

    if key == "follow_up_readiness":
        if metrics["follow_up_signal"] and metrics["machine_feeds_ok"]:
            return result("PASS", 4, f"Sample PDPs exposed {metrics['follow_up_signal_count']} follow-up-answer cues alongside structured feed/schema signals.", "Maintain public data depth for follow-up-answer use cases.", sources="product samples+feeds")
        if metrics["follow_up_signal"] or metrics["machine_feeds_ok"]:
            return result("PARTIAL", 2, f"follow_up_signals={metrics['follow_up_signal_count']}, machine_feeds={metrics['machine_feeds_ok']}", "Add richer FAQ/spec/support content or strengthen structured feeds.", sources="product samples+feeds")
        return result("FAIL", 1, "Public storefront signals looked too thin for reliable follow-up answering.", "Increase publicly visible structured facts and support content.", sources="product samples+feeds")

    if key == "related_products":
        if metrics["related_products_any"]:
            return result("PASS", 4, "Sample PDPs showed visible related-product or complementary-product modules.", "Keep related products explicit and relevant to real intent paths.", sources="product samples")
        return result("FAIL", 1, "No strong visible related-product/cross-sell modules were detected on sampled PDPs.", "Add explicit alternatives, complements, or complete-the-look modules where helpful.", sources="product samples")

    if key == "clothing_attribute_depth":
        evidence = f"attribute_categories={', '.join(metrics['clothing_attribute_categories']) or 'none'}"
        if metrics["clothing_attribute_depth_ok"]:
            return result("PASS", 5, evidence, "Keep fabric/fit/weather/style signals visible and structured on apparel PDPs.", sources="product samples")
        if metrics["clothing_attribute_categories"]:
            return result("PARTIAL", 3, evidence, "Expose deeper apparel fields such as fabric, fit, weather, care, and use-case more consistently.", sources="product samples")
        return result("FAIL", 1, evidence, "Sampled PDPs showed very weak apparel-specific field depth.", "Add richer clothing attributes to visible copy and structured outputs.", sources="product samples")

    if key == "llms_txt":
        if metrics["llms_ok"]:
            return result("PASS", 4, "An llms.txt path was reachable.", "Keep llms.txt current if you adopt it as an AI access policy surface.", sources="llms.txt")
        return result("FAIL", 1, "No llms.txt path was reachable at /llms.txt or /.well-known/llms.txt.", "Consider publishing llms.txt if you want explicit AI-consumption guidance.", sources="llms.txt")

    if key == "ai_bot_policy":
        if not metrics["ai_bot_blocks"]:
            return result("PASS", 4, "No explicit blocks were detected for the major AI bot list we probed.", "Review robots rules carefully whenever AI-discovery policy changes.", sources="robots.txt")
        return result("FAIL", 1, f"Blocked AI bots detected: {', '.join(metrics['ai_bot_blocks'])}.", "Decide whether these blocks are intentional for AI discovery strategy.", sources="robots.txt")

    if key == "machine_feeds":
        if metrics["machine_feeds_ok"]:
            return result("PASS", 5, "Both Product schema and products.json were available.", "Keep both schema and feed outputs aligned.", sources="product schema+products.json")
        if metrics["products_json_ok"] or metrics["product_schema_all"]:
            return result("PARTIAL", 3, "Only one structured channel (schema or feed) was consistently available.", "Strengthen the missing structured data channel.", sources="product schema+products.json")
        return result("FAIL", 1, "Neither strong Product schema coverage nor products.json was confirmed.", "Expose more machine-readable product data.", sources="product schema+products.json")

    if key == "schema_alignment":
        if metrics["schema_alignment_ok"]:
            return result("PASS", 4, "Sampled PDPs showed Product schema, stable canonicals, and rendered metadata aligned with visible body copy.", "Keep machine-readable and visible product facts aligned.", sources="product samples")
        if metrics["product_schema_all"] or metrics["product_meta_ok"]:
            return result("PARTIAL", 2, "Schema existed but alignment signals were incomplete or inconsistent across rendered title/meta/body content.", "Review schema output against visible PDP content and canonical tags.", sources="product samples")
        return result("FAIL", 1, "Schema alignment could not be trusted from sampled PDPs.", "Audit PDP schema against visible content.", sources="product samples")

    if key == "regional_signals":
        if metrics["regional_signal"]:
            return result("PASS", 4, "Detected hreflang, currency, country-code, or region-selector signals in the public storefront.", "Keep locale/currency signals aligned to live markets.", sources="homepage+content samples")
        return result("PARTIAL", 2, "No strong locale or currency signals were detected publicly.", "Add or validate regionalisation signals if the storefront serves multiple markets.", sources="homepage")

    if key == "mcp":
        if metrics["mcp_ok"]:
            endpoint = metrics.get("mcp_endpoint") or "A public MCP path"
            return result("PASS", 4, f"Detected Shopify Storefront MCP at {endpoint}.", "Use Shopify's native MCP as the base commerce layer for agent-driven catalog, cart, and policy flows.", sources="MCP paths")
        return result("FAIL", 0, "No public MCP manifest/endpoint was detected on the probed paths.", "Only add MCP if there is a real machine-to-machine use case.", sources="MCP paths")

    if key == "mcp_depth":
        evidence = f"endpoint={metrics.get('mcp_endpoint') or 'n/a'}, manifest_tools={metrics['mcp_tool_count']}, rpc_ok={metrics['mcp_rpc_ok']}"
        if metrics.get("mcp_protocol_version"):
            evidence = f"{evidence}, protocol={metrics['mcp_protocol_version']}"
        if metrics["mcp_rpc_ok"] and metrics["mcp_tool_count"] > 0:
            return result("PASS", 5, evidence, "Keep the Shopify MCP tool contract documented and stable for agent flows.", sources="MCP manifest+RPC")
        if metrics["mcp_ok"] or metrics["mcp_rpc_ok"]:
            return result("PARTIAL", 2, evidence, "Finish validating the MCP tool list and RPC transport before relying on it in production agent flows.", sources="MCP manifest+RPC")
        return result("FAIL", 0, evidence, "No usable MCP manifest/RPC depth was detected.", "Add MCP only when you have a concrete machine-to-machine workflow and can support proper RPC.", sources="MCP manifest+RPC")

    raise KeyError(f"Unknown check key: {key}")


def overall_score(results: List[Dict[str, Any]]) -> float:
    scored = [item["score"] for item in results if item["status"] not in {"MANUAL", "OUT OF SCOPE"}]
    return round(sum(scored) / len(scored), 2) if scored else 0.0


def score_for_keys(results_map: Dict[str, Dict[str, Any]], keys: List[str]) -> float:
    values = [results_map[key]["score"] for key in keys if key in results_map and results_map[key]["status"] not in {"MANUAL", "OUT OF SCOPE"}]
    return round(sum(values) / len(values), 2) if values else 0.0


def join_labels(labels: List[str]) -> str:
    if not labels:
        return "None"
    if len(labels) == len(SITE_LABELS):
        return f"All {len(SITE_LABELS)}"
    return ", ".join(labels)


def impacted_labels(site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]], keys: List[str]) -> List[str]:
    labels = []
    for site, site_map in zip(site_results, site_result_maps):
        if any(site_map[key]["status"] != "PASS" for key in keys):
            labels.append(site_label(site))
    return labels


def strongest_site(site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]]) -> str:
    scored = [
        (site_label(site), overall_score(list(site_map.values())))
        for site, site_map in zip(site_results, site_result_maps)
    ]
    scored.sort(key=lambda item: (-item[1], item[0]))
    return scored[0][0]


def build_summary_content(site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]]) -> Dict[str, Any]:
    metrics_by_site = [build_metrics(site) for site in site_results]
    all_sites_label = f"All {len(site_results)}"

    crawl_ready = all(metric["homepage_ok"] and metric["robots_ok"] and metric["sitemap_ok"] for metric in metrics_by_site)
    feeds_ready = all(metric["products_json_ok"] for metric in metrics_by_site)
    summary_parts = []
    if crawl_ready and feeds_ready:
        summary_parts.append(
            f"All {len(site_results)} storefronts have solid crawl foundations with live homepages, robots.txt, sitemap.xml, and products.json."
        )
    elif crawl_ready:
        summary_parts.append(
            "Crawl foundations are broadly in place, but structured feeds are not consistently exposed across every storefront."
        )
    else:
        summary_parts.append(
            "The storefronts need baseline crawlability cleanup before deeper AI-readiness improvements will fully matter."
        )

    summary_parts.append(
        "The main gap is semantic depth rather than discovery: product and collection schema are still shallow, sampled PDPs are thin on follow-up-answer content, and apparel-specific attributes are inconsistently exposed."
    )

    if all(site_map["llms_txt"]["status"] != "PASS" for site_map in site_result_maps) and all(site_map["mcp"]["status"] != "PASS" for site_map in site_result_maps):
        summary_parts.append(
            "None of the audited sites currently expose an explicit AI access layer such as llms.txt or a usable public MCP surface."
        )
    elif all(site_map["mcp"]["status"] == "PASS" for site_map in site_result_maps):
        summary_parts.append(
            f"All {len(site_results)} storefronts already expose Shopify's native Storefront MCP, so the MCP conversation shifts from endpoint availability to whether a separate custom MCP layer is needed for brand-specific selling logic."
        )

    summary_parts.append(
        f"{strongest_site(site_results, site_result_maps)} is currently the strongest on public storefront structure, but all sites still need deeper machine-readable commerce data for intent-driven retrieval."
    )

    highlights = []
    if crawl_ready:
        highlights.append(
            (
                "Crawl foundations are in place: homepage, robots.txt, sitemap.xml, and products.json are available.",
                all_sites_label,
                "Technical discovery is not the main blocker.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["product_schema_presence", "product_schema_completeness", "schema_depth"])
    if labels:
        highlights.append(
            (
                "Product schema depth is weak overall across sampled PDPs.",
                join_labels(labels),
                "AI systems need richer machine-readable commerce fields to rank and recommend accurately.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["collection_metadata", "collection_content", "schema_depth"])
    if labels:
        highlights.append(
            (
                "Collection and listing semantics are weak, with limited category copy or machine-readable list depth.",
                join_labels(labels),
                "Category intent stays under-explained without stronger collection semantics.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["clothing_attribute_depth"])
    if labels:
        highlights.append(
            (
                "Apparel-specific field depth is thin across sampled products.",
                join_labels(labels),
                "Fabric, fit, weather, care, and use-case signals are critical for high-intent apparel queries.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["semantic_depth", "follow_up_readiness"])
    if labels:
        highlights.append(
            (
                "Sampled PDPs are generally thin on FAQ, sizing, reviews, or other follow-up-answer support content.",
                join_labels(labels),
                "Thin PDP support content weakens AI answers to fit, suitability, and comparison questions.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["related_products"])
    if labels:
        highlights.append(
            (
                "Visible related-product and alternative-product modules are inconsistent.",
                join_labels(labels),
                "Explicit product relationships improve recommendation quality and cross-sell logic.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["homepage_schema", "homepage_schema_depth"])
    if labels:
        highlights.append(
            (
                "Homepage entity schema is present but shallow or inconsistent.",
                join_labels(labels),
                "Organization, WebSite, and SearchAction signals strengthen brand/entity understanding.",
            )
        )

    labels = impacted_labels(site_results, site_result_maps, ["llms_txt"])
    if labels:
        highlights.append(
            (
                "llms.txt is still absent across the public storefronts.",
                join_labels(labels),
                "The stores now have native Shopify MCP, but they still do not publish an explicit AI-consumption policy surface.",
            )
        )

    actions = build_consolidated_actions(site_results, site_result_maps)

    return {
        "executive_summary": " ".join(summary_parts),
        "highlights": highlights[:8],
        "actions": actions,
        "effort_summary": consolidated_action_effort_summary(actions),
    }


def build_consolidated_actions(site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]]) -> List[Dict[str, Any]]:
    all_sites_label = f"All {len(site_results)}"
    baseline_sites = 3
    total_sites = max(1, len(site_results))

    def scaled(low: float, high: float, scope_sites: int | None = None) -> tuple[float, float]:
        scope = max(1, scope_sites or total_sites)
        factor = scope / baseline_sites
        return (low * factor, high * factor)

    def round_half(value: float) -> float:
        return round(value * 2.0) / 2.0

    def fmt_days(value: float) -> str:
        if float(value).is_integer():
            return str(int(value))
        return f"{value:.1f}".rstrip("0").rstrip(".")

    def action_item(
        key: str,
        action: str,
        domains: str,
        priority: str,
        dev_low: float,
        dev_high: float,
        qa_low: float,
        qa_high: float,
        expected_outcome: str,
        bucket: str,
        scope_sites: int | None = None,
    ) -> Dict[str, Any]:
        dev_low, dev_high = scaled(dev_low, dev_high, scope_sites)
        qa_low, qa_high = scaled(qa_low, qa_high, scope_sites)
        dev_days = round_half((dev_low + dev_high) / 2.0)
        qa_days = max(0.5, round_half((qa_low + qa_high) / 2.0))
        estimate_days = round_half(dev_days + qa_days)

        return {
            "key": key,
            "action": action,
            "domains": domains,
            "priority": priority,
            "estimate": f"{fmt_days(estimate_days)} mandays approx",
            "estimate_basis": f"Dev {fmt_days(dev_days)} + QA {fmt_days(qa_days)}",
            "estimate_days": estimate_days,
            "site_count": total_sites,
            "expected_outcome": expected_outcome,
            "bucket": bucket,
        }

    actions = [
        action_item(
            "deep_pdp_schema",
            "Upgrade PDP schema from shallow presence to deep commerce schema, including brand, offers, category, material, color, size, audience, and additionalProperty fields.",
            all_sites_label,
            "High",
            6.0,
            9.0,
            2.0,
            3.0,
            "Better intent matching, richer retrieval, and stronger product understanding.",
            "core",
            len(site_results),
        ),
        action_item(
            "collection_semantics",
            "Add CollectionPage or ItemList semantics plus stronger category copy and canonical coverage on listing pages.",
            all_sites_label,
            "High",
            4.0,
            6.0,
            2.0,
            2.0,
            "Cleaner category discovery for non-branded and comparison-style queries.",
            "core",
            len(site_results),
        ),
        action_item(
            "apparel_fields",
            "Add explicit apparel field blocks on PDPs for fabric, fit, weather or use-case, care, and style instead of relying on generic marketing copy.",
            all_sites_label,
            "High",
            7.0,
            9.0,
            2.0,
            3.0,
            "Stronger recommendation quality for clothing-specific queries.",
            "core",
            len(site_results),
        ),
        action_item(
            "pdp_faq_support",
            "Add FAQ and support content blocks on PDPs covering sizing, suitability, care, shipping or returns, and product-specific buyer questions.",
            all_sites_label,
            "High",
            7.0,
            9.0,
            2.0,
            3.0,
            "Better follow-up-answer coverage and fewer guessed answers.",
            "core",
            len(site_results),
        ),
        action_item(
            "related_products",
            "Add visible related-products, alternatives, or complete-the-look modules where they are missing or weak.",
            all_sites_label,
            "Medium",
            5.0,
            7.0,
            1.0,
            2.0,
            "Stronger recommendation pathways and better cross-sell teaching signals.",
            "core",
            len(site_results),
        ),
        action_item(
            "homepage_entity_schema",
            "Deepen homepage brand and site entity schema with Organization, WebSite, and SearchAction fields where missing or shallow.",
            all_sites_label,
            "Medium",
            3.0,
            4.0,
            1.0,
            2.0,
            "Improved homepage machine readability and brand framing.",
            "core",
            len(site_results),
        ),
        action_item(
            "native_mcp",
            "Keep Shopify's native Storefront MCP as the base commerce layer and document which shopper flows should use it.",
            all_sites_label,
            "Medium",
            1.5,
            2.0,
            0.5,
            1.0,
            "Cleaner agent integration without rebuilding standard commerce tools.",
            "core",
            len(site_results),
        ),
        action_item(
            "custom_mcp",
            "If AI-led shopping is strategic, add a separate custom MCP for comparison, sizing and fit advice, recommendation reasons, brand facts, and region-specific guidance because Shopify's native MCP is fixed and not editable.",
            all_sites_label,
            "Medium",
            54.0,
            58.0,
            13.0,
            15.0,
            "Adds the editable advisory layer that Shopify's native MCP does not provide by default.",
            "custom_mcp",
            len(site_results),
        ),
    ]

    # Keep explicitly approved breakup only for the 4-site plan.
    if total_sites == 4:
        forced_breakups = {
            "deep_pdp_schema": (14, 11, 3),
            "collection_semantics": (11, 8, 3),
            "apparel_fields": (15, 12, 3),
            "pdp_faq_support": (15, 12, 3),
            "related_products": (9, 7, 2),
            "homepage_entity_schema": (8, 6, 2),
            "native_mcp": (3, 2, 1),
            "custom_mcp": (85, 68, 17),
        }
        for item in actions:
            if item["key"] in forced_breakups:
                estimate_days, dev_days, qa_days = forced_breakups[item["key"]]
                item["estimate"] = f"{estimate_days} mandays approx"
                item["estimate_basis"] = f"Dev {dev_days} + QA {qa_days}"
                item["estimate_days"] = estimate_days

    return actions


def consolidated_action_effort_summary(actions: List[Dict[str, Any]]) -> List[Tuple[str, str, str]]:
    def fmt_days(value: float) -> str:
        if float(value).is_integer():
            return str(int(value))
        return f"{value:.1f}".rstrip("0").rstrip(".")

    site_count = max((int(item.get("site_count", 0)) for item in actions), default=0)
    if site_count == 4:
        core_days = 75
        custom_days = 85
        total_days = 160
    else:
        core_days = sum(float(item.get("estimate_days", 0)) for item in actions if item.get("bucket") == "core")
        custom_days = sum(float(item.get("estimate_days", 0)) for item in actions if item.get("bucket") == "custom_mcp")
        other_days = sum(float(item.get("estimate_days", 0)) for item in actions if item.get("bucket") not in {"core", "custom_mcp"})
        total_days = core_days + custom_days + other_days

    return [
        ("Core storefront-readiness effort", f"{fmt_days(core_days)} mandays approx", "Includes implementation plus QA or re-test time across the storefront workstreams."),
        ("Custom MCP effort", f"{fmt_days(custom_days)} mandays approx", "Includes custom MCP design, implementation, QA, and pilot validation."),
        ("Total effort including custom MCP", f"{fmt_days(total_days)} mandays approx", "Combined delivery estimate including implementation and QA or re-test time."),
    ]


def clone_style(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.protection = copy(source.protection)


def clear_range(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def estimate_row_height(values: List[Any], widths: List[float], min_height: float = 24.0) -> float:
    wrapped_lines = 1
    for value, width in zip(values, widths):
        text = str(value or "")
        if not text:
            continue
        per_line = max(8, int(width * 0.9))
        line_count = 0
        for chunk in text.splitlines() or [""]:
            line_count += max(1, math.ceil(len(chunk) / per_line))
        wrapped_lines = max(wrapped_lines, line_count)
    return max(min_height, 15.0 * wrapped_lines + 6.0)


def write_summary_sheet(ws, site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]], generated_at: str) -> None:
    content = build_summary_content(site_results, site_result_maps)
    thin_side = Side(style="thin", color="D1D5DB")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    title_fill = PatternFill("solid", fgColor="0F172A")
    section_fill = PatternFill("solid", fgColor="1D4ED8")
    header_fill = PatternFill("solid", fgColor="DBEAFE")
    body_fill = PatternFill("solid", fgColor="F8FAFC")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 34

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Shopify Storefront Audit Summary"
    title_cell.font = Font(name="Aptos", size=16, bold=True, color="FFFFFF")
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = border

    ws.merge_cells("A2:F2")
    subtitle = ws["A2"]
    subtitle.value = f"Generated {generated_at} for {', '.join(site_label(site) for site in site_results)}"
    subtitle.font = Font(name="Aptos", size=10, color="334155")
    subtitle.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A4:F4")
    section = ws["A4"]
    section.value = "Executive Summary"
    section.font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    section.fill = section_fill
    section.alignment = Alignment(horizontal="left", vertical="center")
    section.border = border

    ws.merge_cells("A5:F7")
    summary_cell = ws["A5"]
    summary_cell.value = content["executive_summary"]
    summary_cell.font = Font(name="Aptos", size=11, color="0F172A")
    summary_cell.fill = body_fill
    summary_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    summary_cell.border = border
    for row_idx in range(5, 8):
        ws.row_dimensions[row_idx].height = 24

    ws.merge_cells("A9:F9")
    section = ws["A9"]
    section.value = "Finding Highlights"
    section.font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    section.fill = section_fill
    section.alignment = Alignment(horizontal="left", vertical="center")
    section.border = border

    headers = ["#", "Highlight", "Domains", "Why It Matters"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(10, col, value)
        cell.font = Font(name="Aptos", size=10, bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    highlight_row = 11
    for idx, (highlight, domains, why) in enumerate(content["highlights"], start=1):
        values = [idx, highlight, domains, why]
        for col, value in enumerate(values, 1):
            cell = ws.cell(highlight_row, col, value)
            cell.font = Font(name="Aptos", size=10, color="0F172A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = body_fill
            cell.border = border
        highlight_row += 1

    effort_header_row = highlight_row + 2
    ws.merge_cells(start_row=effort_header_row, start_column=1, end_row=effort_header_row, end_column=6)
    section = ws.cell(effort_header_row, 1)
    section.value = "Estimated Delivery Effort"
    section.font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    section.fill = section_fill
    section.alignment = Alignment(horizontal="left", vertical="center")
    section.border = border

    effort_headers = ["#", "Metric", "Estimate", "Estimate Basis"]
    for col, value in enumerate(effort_headers, 1):
        cell = ws.cell(effort_header_row + 1, col, value)
        cell.font = Font(name="Aptos", size=10, bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    effort_row = effort_header_row + 2
    for idx, (metric, estimate, basis) in enumerate(content["effort_summary"], start=1):
        for col, value in enumerate([idx, metric, estimate, basis], 1):
            cell = ws.cell(effort_row, col, value)
            cell.font = Font(name="Aptos", size=10, color="0F172A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = body_fill
            cell.border = border
        ws.row_dimensions[effort_row].height = estimate_row_height(
            [idx, metric, estimate, basis],
            [ws.column_dimensions["A"].width, ws.column_dimensions["B"].width, ws.column_dimensions["C"].width, ws.column_dimensions["D"].width],
            24.0,
        )
        effort_row += 1

    action_header_row = effort_row + 2
    ws.merge_cells(start_row=action_header_row, start_column=1, end_row=action_header_row, end_column=6)
    section = ws.cell(action_header_row, 1)
    section.value = "Action Suggestions"
    section.font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    section.fill = section_fill
    section.alignment = Alignment(horizontal="left", vertical="center")
    section.border = border

    action_columns = ["#", "Action", "Domains", "Priority", "Estimate", "Expected Outcome"]
    for col, value in enumerate(action_columns, 1):
        cell = ws.cell(action_header_row + 1, col, value)
        cell.font = Font(name="Aptos", size=10, bold=True, color="0F172A")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    action_row = action_header_row + 2
    for idx, item in enumerate(content["actions"], start=1):
        values = [idx, item["action"], item["domains"], item["priority"], f'{item["estimate"]} ({item["estimate_basis"]})', item["expected_outcome"]]
        for col, value in enumerate(values, 1):
            cell = ws.cell(action_row, col, value)
            cell.font = Font(name="Aptos", size=10, color="0F172A")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.fill = body_fill
            cell.border = border
        ws.row_dimensions[action_row].height = estimate_row_height(
            values,
            [
                ws.column_dimensions["A"].width,
                ws.column_dimensions["B"].width,
                ws.column_dimensions["C"].width,
                ws.column_dimensions["D"].width,
                ws.column_dimensions["E"].width,
                ws.column_dimensions["F"].width,
            ],
            30.0,
        )
        action_row += 1


def write_readiness_sheet(ws, site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]]) -> None:
    clear_range(ws, 1, 120, 1, 21)
    headers = [
        "",
        "Stage",
        "Domain",
        "What We Check",
        "Shopify Area",
        "Why It Matters for AI Search",
        "Trailberg Status",
        "Trailberg Score",
        "Trailberg Evidence",
        "Lorenzo Status",
        "Lorenzo Score",
        "Lorenzo Evidence",
        "Dream Is Free Status",
        "Dream Is Free Score",
        "Dream Is Free Evidence",
        "Aari Clothing Status",
        "Aari Clothing Score",
        "Aari Clothing Evidence",
    ]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        template = ws.cell(1, 2 if col > 6 else col)
        clone_style(template, cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 4,
        "B": 22,
        "C": 20,
        "D": 34,
        "E": 26,
        "F": 34,
        "G": 16,
        "H": 12,
        "I": 34,
        "J": 16,
        "K": 12,
        "L": 34,
        "M": 18,
        "N": 12,
        "O": 34,
        "P": 18,
        "Q": 12,
        "R": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    mapping = {
        "homepage_foundation": ["homepage_status", "homepage_metadata", "homepage_schema"],
        "crawl_discovery": ["robots", "sitemap"],
        "url_structure": ["url_structure"],
        "collection_foundation": ["collections_exist", "collection_metadata", "collection_content"],
        "product_metadata": ["product_metadata"],
        "product_schema": ["product_schema_presence", "product_schema_completeness"],
        "machine_feeds": ["products_json", "machine_feeds"],
        "agent_context": ["semantic_depth", "follow_up_readiness"],
        "breadcrumbs": ["breadcrumbs"],
        "search_and_facets": ["search_facets"],
        "trust_content": ["trust_content"],
        "llms_txt": ["llms_txt"],
        "ai_bot_policy": ["ai_bot_policy"],
        "mcp_discovery": ["mcp"],
        "regional_signals": ["regional_signals"],
        "out_of_scope": [],
    }

    def summarize(site_map: Dict[str, Dict[str, Any]], summary_key: str) -> Dict[str, Any]:
        if summary_key == "out_of_scope":
            return result("OUT-OF-SCOPE", 0, "Handled in the in-repo brand visibility report workflow.", "Run the brand visibility report generator before producing the final master workbook.", observability="Integrated visibility workflow", sources="Brand visibility report")
        keys = mapping[summary_key]
        subset = [site_map[key] for key in keys]
        score = round(sum(item["score"] for item in subset) / len(subset), 2) if subset else 0
        if all(item["status"] == "PASS" for item in subset):
            status = "PASS"
        elif any(item["status"] == "FAIL" for item in subset):
            status = "PARTIAL"
        else:
            status = "PARTIAL"
        evidence = "; ".join(item["evidence"] for item in subset[:2])
        recommendation = "; ".join(item["recommendation"] for item in subset[:2])
        return result(status, int(round(score)), evidence, recommendation)

    row = 2
    for stage, domain, what_check, area, why, summary_key in SUMMARY_ROWS:
        values = [stage, domain, what_check, area, why]
        for idx, value in enumerate(values, 2):
            cell = ws.cell(row, idx, value)
            template = ws.cell(2, min(idx, 6))
            clone_style(template, cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        col = 7
        for site_map in site_result_maps:
            site_summary = summarize(site_map, summary_key)
            cells = [
                site_summary["status"],
                site_summary["score"],
                site_summary["evidence"],
            ]
            for offset, value in enumerate(cells):
                target = ws.cell(row, col + offset, value)
                template = ws.cell(2, 6 if offset != 1 else 6)
                clone_style(ws.cell(2, 6), target)
                target.alignment = Alignment(vertical="top", wrap_text=True)
            col += 3
        row += 1

    ws.freeze_panes = "A2"


def write_detail_sheet(ws, site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]]) -> None:
    clear_range(ws, 1, 220, 1, 24)
    headers = [
        "",
        "Phase",
        "Team",
        "Audit Domain",
        "What to Check",
        "Trailberg Status",
        "Trailberg Score",
        "Trailberg Evidence / Notes",
        "Lorenzo Status",
        "Lorenzo Score",
        "Lorenzo Evidence / Notes",
        "Dream Is Free Status",
        "Dream Is Free Score",
        "Dream Is Free Evidence / Notes",
        "Aari Clothing Status",
        "Aari Clothing Score",
        "Aari Clothing Evidence / Notes",
        "Why This Matters for AI Agents",
        "Recommended Action",
        "Observability",
        "Source Signals",
    ]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        clone_style(ws.cell(1, 2 if col > 18 else col), cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = {
        "A": 3.5,
        "B": 19,
        "C": 13,
        "D": 21,
        "E": 46,
        "F": 16,
        "G": 12,
        "H": 34,
        "I": 16,
        "J": 12,
        "K": 34,
        "L": 18,
        "M": 12,
        "N": 34,
        "O": 32,
        "P": 32,
        "Q": 18,
        "R": 18,
        "S": 32,
        "T": 32,
        "U": 18,
        "V": 18,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    row = 2
    for check in DETAIL_CHECKS:
        ws.cell(row, 2, check["phase"])
        ws.cell(row, 3, check["team"])
        ws.cell(row, 4, check["domain"])
        ws.cell(row, 5, check["question"])
        for col in range(2, 6):
            clone_style(ws.cell(2, col), ws.cell(row, col))
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        col = 6
        recommendations = []
        observability = []
        sources = []
        for site_map in site_result_maps:
            res = site_map[check["key"]]
            cells = [res["status"], res["score"], res["evidence"]]
            for offset, value in enumerate(cells):
                target = ws.cell(row, col + offset, value)
                source_style = ws.cell(2, 8 if offset == 2 else 8)
                clone_style(ws.cell(2, 8), target)
                target.alignment = Alignment(vertical="top", wrap_text=True)
            recommendations.append(res["recommendation"])
            observability.append(res["observability"])
            sources.append(res["sources"])
            col += 3
        summary_col = 6 + (3 * len(site_result_maps))
        ws.cell(row, summary_col, check["why"])
        ws.cell(row, summary_col + 1, " | ".join(dict.fromkeys(recommendations)))
        ws.cell(row, summary_col + 2, " | ".join(dict.fromkeys(observability)))
        ws.cell(row, summary_col + 3, " | ".join(dict.fromkeys(sources)))
        for col in range(summary_col, summary_col + 4):
            clone_style(ws.cell(2, 10), ws.cell(row, col))
            ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    ws.freeze_panes = "A2"


def write_overall_sheet(ws, site_results: List[Dict[str, Any]], site_result_maps: List[Dict[str, Dict[str, Any]]]) -> None:
    clear_range(ws, 1, 40, 1, 36)
    blocks = [(2 + (idx * 7), 7 + (idx * 7)) for idx in range(len(site_results))]
    metrics_labels = [
        "Homepage Status",
        "robots.txt",
        "Sitemap URLs",
        "Collection URLs",
        "Product URLs",
        "products.json",
        "llms.txt",
        "Public MCP",
        "Overall Public Score",
        "Structured Data Score",
        "Content Semantics Score",
        "Crawlability Score",
    ]

    for (start_col, _), site, site_map in zip(blocks, site_results, site_result_maps):
        label = site_label(site)
        metrics = build_metrics(site)
        summary = {
            "Homepage Status": f"{site['homepage'].get('status')} / {'indexable' if metrics['homepage_indexable'] else 'review'}",
            "robots.txt": "Yes" if metrics["robots_ok"] else "No",
            "Sitemap URLs": site["sitemap"].get("total_urls", 0),
            "Collection URLs": metrics["collection_count"],
            "Product URLs": metrics["product_count"],
            "products.json": "Yes" if metrics["products_json_ok"] else "No",
            "llms.txt": "Yes" if metrics["llms_ok"] else "No",
            "Public MCP": "Yes" if metrics["mcp_ok"] else "No",
            "Overall Public Score": overall_score(list(site_map.values())),
            "Structured Data Score": score_for_keys(site_map, ["homepage_schema", "product_schema_presence", "product_schema_completeness", "machine_feeds", "schema_alignment"]),
            "Content Semantics Score": score_for_keys(site_map, ["collection_content", "semantic_depth", "trust_content", "follow_up_readiness"]),
            "Crawlability Score": score_for_keys(site_map, ["homepage_status", "robots", "sitemap", "url_structure", "collection_metadata", "product_metadata"]),
        }
        ws.cell(2, start_col, site["base_url"])
        ws.cell(3, start_col, "Metric")
        ws.cell(3, start_col + 1, label)
        for col in (start_col, start_col + 1):
            clone_style(ws.cell(3, 2), ws.cell(3, col))
        clone_style(ws.cell(2, 2), ws.cell(2, start_col))
        for row_offset, metric_name in enumerate(metrics_labels, start=4):
            ws.cell(row_offset, start_col, metric_name)
            ws.cell(row_offset, start_col + 1, summary[metric_name])
            clone_style(ws.cell(4, 2), ws.cell(row_offset, start_col))
            clone_style(ws.cell(4, 3), ws.cell(row_offset, start_col + 1))
            ws.cell(row_offset, start_col).alignment = Alignment(vertical="top", wrap_text=True)
            ws.cell(row_offset, start_col + 1).alignment = Alignment(vertical="top", wrap_text=True)


def build_page_findings(site: Dict[str, Any], site_map: Dict[str, Dict[str, Any]]) -> List[List[Any]]:
    metrics = build_metrics(site)
    findings = [
        ["Sitewide", site["base_url"], "Homepage foundation", site_map["homepage_status"]["status"], site_map["homepage_status"]["score"], site_map["homepage_status"]["evidence"], site_map["homepage_status"]["recommendation"], "homepage"],
        ["Sitewide", f"{site['base_url']}/robots.txt", "robots.txt + sitemap declaration", site_map["robots"]["status"], site_map["robots"]["score"], site_map["robots"]["evidence"], site_map["robots"]["recommendation"], "robots.txt"],
        ["Sitewide", f"{site['base_url']}/sitemap.xml", "Sitemap coverage", site_map["sitemap"]["status"], site_map["sitemap"]["score"], site_map["sitemap"]["evidence"], site_map["sitemap"]["recommendation"], "sitemap.xml"],
        ["Sitewide", f"{site['base_url']}/products.json?limit=250", "products.json feed", site_map["products_json"]["status"], site_map["products_json"]["score"], site_map["products_json"]["evidence"], site_map["products_json"]["recommendation"], "products.json"],
        ["Sitewide", f"{site['base_url']}/llms.txt", "llms.txt", site_map["llms_txt"]["status"], site_map["llms_txt"]["score"], site_map["llms_txt"]["evidence"], site_map["llms_txt"]["recommendation"], "llms.txt"],
        ["Sitewide", metrics.get("mcp_endpoint") or f"{site['base_url']}/api/mcp", "Native Shopify MCP", site_map["mcp"]["status"], site_map["mcp"]["score"], site_map["mcp"]["evidence"], site_map["mcp"]["recommendation"], "MCP paths"],
    ]

    for page in site["samples"]["collection_pages"][:2]:
        findings.append(
            [
                "Collection",
                page.get("final_url") or page.get("requested_url"),
                "Collection metadata",
                "PASS" if page.get("h1_count", 0) >= 1 and page.get("canonical_matches") else "PARTIAL",
                5 if page.get("h1_count", 0) >= 1 and page.get("canonical_matches") else 2,
                f"title_len={page.get('title_length')}, meta_len={page.get('meta_description_length')}, words={page.get('word_count')}",
                "Strengthen collection intro copy and metadata where thin.",
                "collection sample",
            ]
        )

    for page in site["samples"]["product_pages"][:3]:
        findings.append(
            [
                "Product",
                page.get("final_url") or page.get("requested_url"),
                "PDP structured data",
                "PASS" if page.get("product_schema_count", 0) > 0 else "FAIL",
                5 if page.get("product_schema_count", 0) > 0 else 1,
                f"schema_types={', '.join(page.get('json_ld_types', [])[:6])}, breadcrumb_count={page.get('breadcrumb_schema_count')}, missing_alt={page.get('images_without_alt')}",
                "Add or harden Product/Breadcrumb schema and clean up missing alt text where needed.",
                "product sample",
            ]
        )
        findings.append(
            [
                "Product",
                page.get("final_url") or page.get("requested_url"),
                "PDP content depth",
                "PASS" if page.get("word_count", 0) >= 180 and (page.get("mentions_faq") or page.get("mentions_size_guide") or page.get("mentions_reviews")) else "PARTIAL",
                4 if page.get("word_count", 0) >= 180 and (page.get("mentions_faq") or page.get("mentions_size_guide") or page.get("mentions_reviews")) else 2,
                f"words={page.get('word_count')}, faq={page.get('mentions_faq')}, size_guide={page.get('mentions_size_guide')}, reviews={page.get('mentions_reviews')}",
                "Add visible support, sizing, or FAQ content to answer follow-up questions.",
                "product sample",
            ]
        )

    return findings


def write_page_sheet(ws, site: Dict[str, Any], site_map: Dict[str, Dict[str, Any]]) -> None:
    clear_range(ws, 1, 220, 1, 12)
    headers = ["", "Page Type", "URL", "Check", "Status", "Score (0-5)", "Evidence", "Recommendation", "Source Signals"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        clone_style(ws.cell(1, 2 if col > 9 else col), cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    widths = {
        "A": 3.25,
        "B": 13,
        "C": 52,
        "D": 28,
        "E": 14,
        "F": 11,
        "G": 42,
        "H": 42,
        "I": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row_idx, row_values in enumerate(build_page_findings(site, site_map), start=2):
        for col, value in enumerate(row_values, 1):
            cell = ws.cell(row_idx, col, value)
            clone_style(ws.cell(2, 2 if col > 9 else col), cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def write_template_sheet(ws) -> None:
    clear_range(ws, 1, 80, 1, 12)
    headers = ["Website", "Page Type", "URL", "Check", "Status", "Score (0-5)", "Evidence", "Recommendation", "Source Signals", "Notes"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(1, col, value)
        clone_style(ws.cell(1, 1 if col == 1 else min(col, 5)), cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.cell(2, 1, "future-domain.com")
    ws.cell(2, 2, "Sitewide")
    ws.cell(2, 3, "https://future-domain.com/")
    ws.cell(2, 4, "Homepage foundation")
    ws.cell(2, 5, "PASS / PARTIAL / FAIL / OUT OF SCOPE")
    ws.cell(2, 6, 0)
    ws.cell(2, 7, "Short factual observation from the crawl.")
    ws.cell(2, 8, "Next action for the team.")
    ws.cell(2, 9, "homepage / robots.txt / sitemap.xml / product sample")
    ws.cell(2, 10, "Optional reviewer notes")
    for col in range(1, 11):
        clone_style(ws.cell(1, min(col, 5)), ws.cell(2, col))
        ws.cell(2, col).alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A": 22, "B": 14, "C": 44, "D": 26, "E": 26, "F": 12, "G": 40, "H": 36, "I": 24, "J": 24}.items():
        ws.column_dimensions[col].width = width


def write_methodology_sheet(ws) -> None:
    clear_range(ws, 1, 60, 1, 8)
    headers = ["", "Category", "What It Tests", "Why It Matters", "What Counts As Pass", "What Is Out of Scope for This Public Audit"]
    for col, value in enumerate(headers, 1):
        cell = ws.cell(2 if col > 1 else 1, col, value)
        source_cell = ws.cell(2, min(max(col, 2), 6))
        clone_style(source_cell, cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    rows = [
        ("Technical crawlability", "Homepage, robots.txt, sitemap.xml, canonicals, indexation", "These are the minimum conditions for external discovery.", "Live 200 pages, indexable markup, robots + sitemap available.", "Hidden application logic or configuration that is not rendered publicly."),
        ("Structured commerce data", "Product JSON-LD, breadcrumb schema, products.json", "This makes product facts machine-readable without brittle scraping.", "Product pages expose stable schema and feed endpoints.", "Hidden source fields or non-public systems that are not rendered in HTML, schema, or public feeds."),
        ("Semantic content depth", "FAQ, support cues, size guides, collection copy", "Agents need follow-up-answer material, not only merchandising copy.", "Sampled pages expose helpful, visible support content.", "Internal documentation or support workflows that are not linked or rendered publicly."),
        ("AI access policy", "llms.txt and AI bot directives", "These are emerging signals for explicit AI consumption policy.", "llms.txt exists and robots does not unintentionally block target AI bots.", "Private licensing terms or unpublished content-rights policy."),
        ("MCP strategy", "Native Shopify MCP and optional custom MCP use cases", "Shopify's native MCP covers baseline commerce and is not editable, so a separate custom MCP is the right path for brand-specific advisory tools.", "Native MCP responds successfully and any custom MCP plans are scoped to real business use cases.", "Private agent orchestration or custom tool contracts that are not publicly exposed."),
        ("Regional context", "hreflang, currencies, market/country signals", "Regional context reduces price/availability hallucination.", "Locale or currency signals are visible in the public storefront.", "Regional logic not visible in rendered locale, currency, hreflang, or public country selectors."),
    ]
    for row_idx, row_values in enumerate(rows, start=3):
        for col, value in enumerate([""] + list(row_values), 1):
            cell = ws.cell(row_idx, col, value)
            clone_style(ws.cell(3, min(max(col, 2), 6)), cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, width in {"A": 4.5, "B": 20, "C": 32, "D": 28, "E": 30, "F": 32}.items():
        ws.column_dimensions[col].width = width


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    template_path = Path(args.template)
    input_path = Path(args.input)
    output_path = Path(args.output)

    with input_path.open("r", encoding="utf-8") as fh:
        payload = json.load(fh)
    preferred_order = ["trailberg.com", "lorenzo.world", "dreamisfree.com", "aariclothing.com"]
    site_results_by_domain = {site["domain"]: site for site in payload["results"]}
    site_results = [site_results_by_domain[domain] for domain in preferred_order if domain in site_results_by_domain]
    site_results.extend(site for site in payload["results"] if site["domain"] not in preferred_order)
    site_result_maps = []
    for site in site_results:
        metrics = build_metrics(site)
        checks = {check["key"]: eval_check(check["key"], metrics) for check in DETAIL_CHECKS}
        site_result_maps.append(checks)

    wb = load_workbook(template_path)

    wb["Lorenzo_Visibility Tracker"].title = "Lorenzo_Page Audit"
    wb["Dream Is Free_Visibility Tracke"].title = "DreamIsFree_Page Audit"
    wb["Trailberg_Visibility Tracker"].title = "Trailberg_Page Audit"
    wb["AARI_Visibility Tracker"].title = "Aari_Page Audit"
    wb["Query Generation"].title = "Audit_Methodology"
    if "Reusable_Template" not in wb.sheetnames:
        wb.create_sheet("Reusable_Template")

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    wb.create_sheet("Summary", 0)

    generated_at = payload.get("generated_at", datetime.utcnow().isoformat())
    write_summary_sheet(wb["Summary"], site_results, site_result_maps, generated_at)
    write_readiness_sheet(wb["Shopify_Agent Readiness"], site_results, site_result_maps)
    write_detail_sheet(wb["Shopify_Internal Structure Audi"], site_results, site_result_maps)
    write_overall_sheet(wb["Overall Reporting "], site_results, site_result_maps)
    page_sheet_map = {
        "trailberg.com": "Trailberg_Page Audit",
        "lorenzo.world": "Lorenzo_Page Audit",
        "dreamisfree.com": "DreamIsFree_Page Audit",
        "aariclothing.com": "Aari_Page Audit",
    }
    for site, site_map in zip(site_results, site_result_maps):
        sheet_name = page_sheet_map.get(site["domain"])
        if sheet_name and sheet_name in wb.sheetnames:
            write_page_sheet(wb[sheet_name], site, site_map)
    write_template_sheet(wb["Reusable_Template"])
    write_methodology_sheet(wb["Audit_Methodology"])

    wb.properties.modified = datetime.utcnow()
    wb.properties.creator = "Codex"
    wb.properties.lastModifiedBy = "Codex"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
