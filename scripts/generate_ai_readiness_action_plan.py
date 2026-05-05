#!/usr/bin/env python3

from __future__ import annotations

import argparse
import importlib.util
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


DEFAULT_CONFIG_PATH = Path("configs/audit_config.json")
DEFAULT_STOREFRONT_AUDIT = Path("output/data/shopify_storefront_audit.json")
DEFAULT_ADMIN_AUDIT = Path("output/data/shopify_admin_audit.json")

PRIORITY_ORDER = {"P0": 0, "P1": 1, "P2": 2}
STATUS_ORDER = {"FAIL": 0, "PARTIAL": 1, "PASS": 2}

PRIORITY_MAP = {
    "homepage_status": "P0",
    "homepage_metadata": "P0",
    "homepage_schema": "P0",
    "homepage_schema_depth": "P0",
    "robots": "P0",
    "sitemap": "P0",
    "url_structure": "P0",
    "collections_exist": "P0",
    "collection_metadata": "P0",
    "product_metadata": "P0",
    "product_schema_presence": "P0",
    "product_schema_completeness": "P0",
    "schema_depth": "P0",
    "schema_alignment": "P0",
    "image_alt": "P0",
    "collection_content": "P1",
    "breadcrumbs": "P1",
    "products_json": "P1",
    "semantic_depth": "P1",
    "search_facets": "P1",
    "trust_content": "P1",
    "follow_up_readiness": "P1",
    "related_products": "P1",
    "clothing_attribute_depth": "P1",
    "ai_bot_policy": "P1",
    "machine_feeds": "P1",
    "regional_signals": "P1",
    "llms_txt": "P2",
    "mcp": "P2",
    "mcp_depth": "P2",
}

OPTIONAL_KEYS = {"llms_txt", "mcp", "mcp_depth"}

ACTION_OVERRIDES = {
    "homepage_metadata": [
        "Enforce one H1 on the homepage and keep title/meta description within the audit target ranges.",
        "Update the theme SEO fields so title, H1, and meta description describe the same brand proposition.",
    ],
    "homepage_schema": [
        "Add Organization and WebSite JSON-LD on the homepage.",
        "Populate brand identity fields consistently, including name, URL, logo, and sameAs when available.",
    ],
    "homepage_schema_depth": [
        "Extend homepage schema with SearchAction and stronger identity cues instead of basic presence only.",
        "Keep structured brand facts synchronized with visible homepage copy.",
    ],
    "product_metadata": [
        "Standardize PDP title, meta description, H1, and canonical generation across templates.",
        "Add publishing QA so PDP metadata does not regress on new products.",
    ],
    "product_schema_presence": [
        "Guarantee Product JSON-LD on every PDP template and variant state.",
        "Validate that schema renders on live pages, not only in Shopify data.",
    ],
    "product_schema_completeness": [
        "Expand Product schema to include brand, description, offers, price/currency, availability, and canonical product URL.",
        "Use the same product facts in visible copy and schema payloads.",
    ],
    "schema_depth": [
        "Add richer commerce semantics such as category, audience/use-case, material/fabric, and collection/list relationships.",
        "Model apparel attributes explicitly instead of leaving them buried in generic copy.",
    ],
    "schema_alignment": [
        "Audit PDP schema against rendered title, description, canonical URL, and on-page product facts.",
        "Remove mismatches between storefront copy and machine-readable values before expanding schema further.",
    ],
    "image_alt": [
        "Add alt-text QA to product and collection publishing workflows.",
        "Ensure gallery and supporting media alts describe the specific product view rather than staying blank.",
        "Audit rendered PDP gallery media so product imagery does not depend on weak or missing alt output.",
    ],
    "breadcrumbs": [
        "Render visible breadcrumbs and BreadcrumbList schema on PDPs and collection pages.",
        "Keep breadcrumb taxonomy aligned with real collection hierarchy.",
    ],
    "semantic_depth": [
        "Add FAQ, sizing, materials, care, shipping, returns, and support cues directly on PDPs.",
        "Prioritize answerable factual content over generic marketing copy.",
    ],
    "trust_content": [
        "Make About, Help, Shipping, Returns, and policy links consistently discoverable in nav or footer.",
        "Reduce dependence on hidden or poorly linked support pages.",
    ],
    "follow_up_readiness": [
        "Publish enough structured facts and support content for follow-up questions about fit, use-case, delivery, and returns.",
        "Reuse PDP modules for FAQs/specs rather than leaving answers in private systems only.",
    ],
    "related_products": [
        "Implement explicit related, alternative, complementary, or complete-the-look modules on PDPs.",
        "Feed those modules from consistent merchandising logic instead of ad hoc manual links.",
    ],
    "clothing_attribute_depth": [
        "Expose apparel attributes such as fabric, fit, weather suitability, care, and use-case consistently on PDPs.",
        "Promote these fields into structured outputs where possible.",
    ],
    "machine_feeds": [
        "Strengthen the second machine-readable channel so schema and feed coverage both exist.",
        "If products.json is insufficient, provide a stable feed or equivalent structured export.",
    ],
    "llms_txt": [
        "Publish `llms.txt` only if you want an explicit AI-consumption policy surface.",
        "Treat this as optional after foundation and structured-data fixes are complete.",
    ],
    "mcp": [
        "Use Shopify's native Storefront MCP as the base commerce layer for catalog search, cart operations, and policy answers.",
        "Document which shopper journeys should call the native MCP before adding any custom tool layer.",
    ],
    "mcp_depth": [
        "Keep the native Shopify MCP contract stable and documented if agents will rely on it.",
        "Only add more MCP surface area when the required shopper workflow is not covered by Shopify's built-in tools.",
    ],
    "custom_mcp": [
        "Design a custom MCP only for workflows Shopify's native MCP does not explain well enough, such as recommendation reasons, product comparison, sizing or fit advice, brand facts, and market-specific guidance.",
        "Keep custom MCP tools tightly scoped to brand-specific advisory logic instead of rebuilding catalog, cart, or policy tools that Shopify already provides.",
    ],
}


def load_config(config_path: Path) -> Dict[str, Any]:
    if not config_path.exists():
        return {"sites": []}
    return json.loads(config_path.read_text(encoding="utf-8"))


def load_existing_module(path: Path):
    spec = importlib.util.spec_from_file_location(path.stem, path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def load_json(path: Path) -> Dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def load_visibility_workbook(path: Path) -> Dict[str, List[Dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_by_sheet: Dict[str, List[Dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        raw_rows = list(sheet.iter_rows(values_only=True))
        if not raw_rows:
            rows_by_sheet[sheet_name] = []
            continue
        headers = [str(value).strip() if value is not None else "" for value in raw_rows[0]]
        parsed_rows: List[Dict[str, Any]] = []
        for row in raw_rows[1:]:
            record: Dict[str, Any] = {}
            has_value = False
            for index, header in enumerate(headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else None
                if value not in (None, ""):
                    has_value = True
                record[header] = value if value is not None else ""
            if has_value:
                parsed_rows.append(record)
        rows_by_sheet[sheet_name] = parsed_rows
    return rows_by_sheet


def label_map(config: Dict[str, Any], fallback_sites: List[Dict[str, Any]]) -> Dict[str, str]:
    sites = config.get("sites", [])
    if sites:
        return {site["domain"]: site["label"] for site in sites}
    return {site["domain"]: site.get("domain", site["domain"]) for site in fallback_sites}


def check_metadata(generator_module) -> Dict[str, Dict[str, str]]:
    meta = {}
    for check in generator_module.DETAIL_CHECKS:
        meta[check["key"]] = {
            "phase": check["phase"],
            "team": check["team"],
            "domain": check["domain"],
            "question": check["question"],
            "why": check["why"],
        }
    return meta


def action_items_for(key: str, recommendation: str) -> List[str]:
    items = ACTION_OVERRIDES.get(key, [])
    if recommendation and recommendation not in items:
        items = items + [recommendation]
    return items[:3]


def safe_float(value: Any) -> float | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def safe_json_list(value: Any) -> List[Any]:
    text = str(value or "").strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        return []
    return parsed if isinstance(parsed, list) else []


def pct(value: float) -> str:
    return f"{round(value * 100)}%"


def build_visibility_findings(
    visibility_rows_by_sheet: Dict[str, List[Dict[str, Any]]] | None,
    config: Dict[str, Any],
    labels: Dict[str, str],
) -> Dict[str, List[Dict[str, Any]]]:
    if not visibility_rows_by_sheet:
        return {}

    findings_by_site: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for site in config.get("sites", []):
        domain = site["domain"]
        site_label = labels.get(domain, site["label"])
        sheet_name = site.get("visibility_report_sheet")
        rows = visibility_rows_by_sheet.get(sheet_name, [])
        if not rows:
            continue

        total = len(rows)
        mentioned_rows = [row for row in rows if str(row.get("Brand Mentioned", "")).strip().upper() == "Y"]
        mention_rate = len(mentioned_rows) / total if total else 0.0
        rank_values = [value for value in (safe_float(row.get("Brand Position")) for row in mentioned_rows) if value is not None]
        avg_rank = (sum(rank_values) / len(rank_values)) if rank_values else None

        target_domain = str(site["domain"]).lower()
        cited_rows = []
        product_cited_rows = []
        for row in rows:
            citations = [str(item).lower() for item in safe_json_list(row.get("citations_json"))]
            has_target_citation = any(target_domain in item for item in citations)
            has_product_citation = any(target_domain in item and "/products/" in item for item in citations)
            if has_target_citation:
                cited_rows.append(row)
            if has_product_citation:
                product_cited_rows.append(row)
        citation_rate = len(cited_rows) / total if total else 0.0
        product_citation_rate = len(product_cited_rows) / total if total else 0.0

        mention_priority = "P0" if mention_rate < 0.2 else "P1"
        mention_evidence = (
            f"{site_label} was mentioned in {len(mentioned_rows)}/{total} visibility prompts ({pct(mention_rate)})."
            + (f" Average mention rank was {avg_rank:.1f}." if avg_rank is not None else " Average rank could not be established because the brand was rarely mentioned.")
        )
        findings_by_site[domain].append(
            {
                "key": "brand_visibility",
                "site": domain,
                "site_label": site_label,
                "status": "FAIL" if mention_rate < 0.2 else "PARTIAL",
                "score": 1 if mention_rate < 0.2 else 2,
                "priority": mention_priority,
                "title": "AI Brand Visibility",
                "phase": "AI Discovery",
                "team": "SEO + Content",
                "area": "Visibility",
                "why": "If the brand is not recalled in category, comparison, and recommendation prompts, agents will default to competitors.",
                "evidence": mention_evidence,
                "recommendation": "Increase branded entity recall with stronger homepage/about identity, clearer brand positioning, and more consistent branded supporting content.",
                "actions": [
                    "Strengthen homepage, About, and help content so the brand proposition is explicit in crawlable copy.",
                    "Align homepage Organization/WebSite schema, on-page claims, and branded internal links so the brand is easier for AI systems to anchor.",
                    "Publish more comparison, use-case, and review-style content that names the brand in the same query patterns used in the visibility workbook.",
                ],
                "optional": False,
            }
        )

        product_priority = "P0" if product_citation_rate == 0 else "P1"
        product_evidence = (
            f"Target-domain citations appeared in {len(cited_rows)}/{total} prompts ({pct(citation_rate)}), "
            f"and direct product citations appeared in {len(product_cited_rows)}/{total} prompts ({pct(product_citation_rate)})."
        )
        findings_by_site[domain].append(
            {
                "key": "product_visibility",
                "site": domain,
                "site_label": site_label,
                "status": "FAIL" if product_citation_rate == 0 else "PARTIAL",
                "score": 1 if product_citation_rate == 0 else 2,
                "priority": product_priority,
                "title": "AI Product Citation Coverage",
                "phase": "AI Discovery",
                "team": "Content + Ecommerce",
                "area": "Visibility",
                "why": "If agents do not cite the site or product URLs, recommendations will rely on weaker third-party descriptions instead of first-party facts.",
                "evidence": product_evidence,
                "recommendation": "Increase first-party citation likelihood by making PDP facts denser, more answerable, and easier to cite.",
                "actions": [
                    "Expand PDP facts, comparisons, FAQs, sizing, materials, and use-case sections so product pages are citation-worthy.",
                    "Ensure product naming, canonical URLs, schema, and visible PDP copy stay tightly aligned across collections and product pages.",
                    "Link editorial, collections, and PDPs together so category prompts can resolve to first-party product URLs rather than only brand-homepage mentions.",
                ],
                "optional": False,
            }
        )

        findings_by_site[domain].sort(key=severity_key)
    return findings_by_site


def severity_key(item: Dict[str, Any]) -> tuple[int, int, str]:
    return (
        PRIORITY_ORDER.get(item["priority"], 99),
        STATUS_ORDER.get(item["status"], 99),
        item["title"].lower(),
    )


def build_site_findings(
    storefront_payload: Dict[str, Any],
    admin_payload: Dict[str, Any] | None,
    generator_module,
    labels: Dict[str, str],
) -> Dict[str, List[Dict[str, Any]]]:
    meta = check_metadata(generator_module)
    admin_map = {site["domain"]: site.get("checks", {}) for site in (admin_payload or {}).get("results", [])}
    findings_by_site: Dict[str, List[Dict[str, Any]]] = defaultdict(list)

    for site in storefront_payload["results"]:
        metrics = generator_module.build_metrics(site)
        checks = {
            check["key"]: generator_module.eval_check(check["key"], metrics)
            for check in generator_module.DETAIL_CHECKS
        }
        for key, result in checks.items():
            if result["status"] not in {"FAIL", "PARTIAL"}:
                continue
            key_meta = meta.get(
                key,
                {
                    "phase": "Operational Readiness",
                    "team": "Platform",
                    "domain": "Admin",
                    "question": key.replace("_", " ").title(),
                    "why": "This depends on platform or operational evidence that is not fully visible in the public storefront render.",
                },
            )
            findings_by_site[site["domain"]].append(
                {
                    "key": key,
                    "site": site["domain"],
                    "site_label": labels.get(site["domain"], site["domain"]),
                    "status": result["status"],
                    "score": result["score"],
                    "priority": PRIORITY_MAP.get(key, "P1"),
                    "title": key_meta["question"],
                    "phase": key_meta["phase"],
                    "team": key_meta["team"],
                    "area": key_meta["domain"],
                    "why": key_meta["why"],
                    "evidence": result["evidence"],
                    "recommendation": result["recommendation"],
                    "actions": action_items_for(key, result["recommendation"]),
                    "optional": key in OPTIONAL_KEYS,
                }
            )
        findings_by_site[site["domain"]].sort(key=severity_key)
    return findings_by_site


def build_cross_site_summary(findings_by_site: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    grouped: Dict[str, Dict[str, Any]] = {}
    for site_findings in findings_by_site.values():
        for finding in site_findings:
            key = finding["key"]
            bucket = grouped.setdefault(
                key,
                {
                    "key": key,
                    "title": finding["title"],
                    "priority": finding["priority"],
                    "team": finding["team"],
                    "phase": finding["phase"],
                    "sites": [],
                    "statuses": [],
                    "actions": finding["actions"],
                    "optional": finding["optional"],
                },
            )
            bucket["sites"].append(finding["site_label"])
            bucket["statuses"].append(f"{finding['site_label']} ({finding['status']})")
    summary = list(grouped.values())
    summary.sort(key=lambda item: (PRIORITY_ORDER.get(item["priority"], 99), len(item["sites"]) * -1, item["title"].lower()))
    return summary


def format_list(items: List[str]) -> str:
    unique_items = []
    for item in items:
        if item not in unique_items:
            unique_items.append(item)
    if not unique_items:
        return ""
    if len(unique_items) == 1:
        return unique_items[0]
    if len(unique_items) == 2:
        return f"{unique_items[0]} and {unique_items[1]}"
    return f"{', '.join(unique_items[:-1])}, and {unique_items[-1]}"


def site_labels(findings_by_site: Dict[str, List[Dict[str, Any]]]) -> List[str]:
    labels: List[str] = []
    for domain, site_findings in findings_by_site.items():
        label = site_findings[0]["site_label"] if site_findings else domain
        if label not in labels:
            labels.append(label)
    return labels


def priority_heading(priority: str) -> str:
    return {
        "P0": "Immediate Priorities",
        "P1": "Next Priorities",
        "P2": "Optional Opportunities",
    }.get(priority, "Other Priorities")


def priority_label(priority: str) -> str:
    return {
        "P0": "Immediate priority",
        "P1": "Next priority",
        "P2": "Optional opportunity",
    }.get(priority, priority)


def effort_range(key: str, priority: str, optional: bool = False) -> tuple[float, float]:
    explicit = {
        "product_visibility": (2.0, 4.0),
        "brand_visibility": (2.0, 4.0),
        "homepage_schema_depth": (2.0, 4.0),
        "schema_depth": (2.0, 4.0),
        "schema_alignment": (2.0, 4.0),
        "machine_feeds": (2.0, 4.0),
        "homepage_schema": (1.0, 2.0),
        "homepage_metadata": (1.0, 2.0),
        "product_schema_presence": (1.0, 2.0),
        "product_schema_completeness": (1.0, 2.0),
        "product_metadata": (1.0, 2.0),
        "breadcrumbs": (1.0, 2.0),
        "semantic_depth": (1.0, 2.0),
        "trust_content": (1.0, 2.0),
        "follow_up_readiness": (1.0, 2.0),
        "related_products": (1.0, 2.0),
        "clothing_attribute_depth": (1.0, 2.0),
        "llms_txt": (0.5, 1.0),
        "custom_mcp": (50.0, 55.0),
    }
    if key in explicit:
        return explicit[key]
    if optional:
        return (0.5, 1.0)
    return {
        "P0": (2.0, 4.0),
        "P1": (1.0, 2.0),
        "P2": (0.5, 1.0),
    }.get(priority, (1.0, 2.0))


def fmt_effort_value(value: float) -> str:
    if value.is_integer():
        return str(int(value))
    return f"{value:.1f}".rstrip("0").rstrip(".")


def effort_estimate(key: str, priority: str, optional: bool = False) -> str:
    low, high = effort_range(key, priority, optional)
    unit = "manday" if low == high == 1 else "mandays"
    return f"{fmt_effort_value(low)}-{fmt_effort_value(high)} {unit} approx"


def total_effort(items: List[Dict[str, Any]]) -> tuple[float, float]:
    low = 0.0
    high = 0.0
    for item in items:
        item_low, item_high = effort_range(item["key"], item["priority"], item.get("optional", False))
        low += item_low
        high += item_high
    return low, high


def table_escape(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def join_html_lines(items: List[str]) -> str:
    return "<br>".join(table_escape(item) for item in items if item)


def normalize_team(team: str) -> str:
    replacements = {
        "Data": "Dev",
        "SEO + Data": "SEO + Dev",
        "Content + Data": "Content + Dev",
    }
    return replacements.get(team, team)


def append_action_table(
    lines: List[str],
    key: str,
    priority: str,
    team: str,
    actions: List[str],
    why: str,
    recommendation: str,
    brand_text: str | None = None,
    evidence: str | None = None,
    optional: bool = False,
) -> None:
    team = normalize_team(team)
    business_items = [f"Why this matters: {why}"]
    if brand_text:
        business_items.append(f"Relevant brands: {brand_text}")
    if evidence:
        business_items.append(f"Current position: {evidence}")
    business_items.append(f"Recommended business move: {recommendation}")

    technical_items = actions[:]
    if optional:
        technical_items.append("Priority note: Optional after the core fixes are stable.")

    lines.append('<table>')
    lines.append('<colgroup>')
    lines.append('<col style="width: 40%;">')
    lines.append('<col style="width: 16%;">')
    lines.append('<col style="width: 44%;">')
    lines.append('</colgroup>')
    lines.append('<thead>')
    lines.append('<tr>')
    lines.append('<th style="vertical-align: top; text-align: left;">Business view</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Lead</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Technical implementation</th>')
    lines.append('</tr>')
    lines.append('</thead>')
    lines.append('<tbody>')
    lines.append('<tr>')
    lines.append(f'<td style="vertical-align: top;">{join_html_lines(business_items)}</td>')
    lines.append(f'<td style="vertical-align: top;">{table_escape(team)}</td>')
    lines.append(f'<td style="vertical-align: top;">{join_html_lines(technical_items)}</td>')
    lines.append('</tr>')
    lines.append('</tbody>')
    lines.append('</table>')
    lines.append("")


def append_consolidated_task_table(lines: List[str], actions: List[Dict[str, Any]]) -> None:
    lines.append('<table>')
    lines.append('<colgroup>')
    lines.append('<col style="width: 5%;">')
    lines.append('<col style="width: 33%;">')
    lines.append('<col style="width: 14%;">')
    lines.append('<col style="width: 10%;">')
    lines.append('<col style="width: 14%;">')
    lines.append('<col style="width: 24%;">')
    lines.append('</colgroup>')
    lines.append('<thead>')
    lines.append('<tr>')
    lines.append('<th style="vertical-align: top; text-align: left;">#</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Action</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Domains</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Priority</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Estimate</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Expected outcome</th>')
    lines.append('</tr>')
    lines.append('</thead>')
    lines.append('<tbody>')
    for idx, item in enumerate(actions, start=1):
        lines.append('<tr>')
        lines.append(f'<td style="vertical-align: top;">{idx}</td>')
        lines.append(f'<td style="vertical-align: top;">{table_escape(item["action"])}</td>')
        lines.append(f'<td style="vertical-align: top;">{table_escape(item["domains"])}</td>')
        lines.append(f'<td style="vertical-align: top;">{table_escape(item["priority"])}</td>')
        lines.append(f'<td style="vertical-align: top;">{table_escape(item["estimate"])} ({table_escape(item["estimate_basis"])})</td>')
        lines.append(f'<td style="vertical-align: top;">{table_escape(item["expected_outcome"])}</td>')
        lines.append('</tr>')
    lines.append('</tbody>')
    lines.append('</table>')
    lines.append("")


def render_markdown(
    findings_by_site: Dict[str, List[Dict[str, Any]]],
    cross_site: List[Dict[str, Any]],
    consolidated_actions: List[Dict[str, Any]],
    effort_summary: List[tuple[str, str, str]],
    generated_at: str,
    visibility_report_path: str | None = None,
    include_custom_mcp_note: bool = False,
) -> str:
    total_sites = len(findings_by_site)
    labels = site_labels(findings_by_site)
    label_text = format_list(labels) if labels else "the audited brands"
    p0_count = sum(1 for item in cross_site if item["priority"] == "P0")
    p1_count = sum(1 for item in cross_site if item["priority"] == "P1")
    workdays_per_week = 5
    project_weeks = 36
    project_capacity = workdays_per_week * project_weeks
    effort_map = {label: value for label, value, _basis in effort_summary}
    core_effort = effort_map.get("Core storefront-readiness effort", "-")
    custom_effort = effort_map.get("Custom MCP effort", "-")
    total_effort_label = effort_map.get("Total effort including custom MCP", core_effort)
    lines: List[str] = []
    lines.append("# AI Readiness Action Plan")
    lines.append("")
    lines.append(f"Generated: {generated_at}")
    lines.append("")
    lines.append("## Executive Summary")
    lines.append("")
    lines.append(f"- Sites reviewed: {total_sites}")
    lines.append(f"- Immediate priorities that need action now: {p0_count}")
    lines.append(f"- Secondary priorities to tackle after the foundations are fixed: {p1_count}")
    lines.append(f"- Planned engagement model: 6 months at {workdays_per_week} working days per week")
    lines.append(f"- Indicative project capacity over that period: {project_capacity} working days")
    lines.append(f"- Estimated core storefront-readiness effort: {core_effort}")
    if include_custom_mcp_note:
        lines.append(f"- Estimated custom MCP effort: {custom_effort}")
        lines.append(f"- Estimated total effort including custom MCP: {total_effort_label}")
    lines.append("- Estimate basis: includes development plus QA or re-test time; it does not try to price stakeholder waiting time, approval delays, or long content-production cycles.")
    if visibility_report_path:
        lines.append(f"- Visibility input workbook: `{visibility_report_path}`")
    lines.append("")
    lines.append("## Consolidated Task List")
    lines.append("")
    lines.append("All task estimates below include implementation plus QA or re-test time.")
    lines.append("")
    append_consolidated_task_table(lines, consolidated_actions)
    lines.append("## Cross-Site Priorities")
    lines.append("")
    for priority in ["P0", "P1", "P2"]:
        section_items = [entry for entry in cross_site if entry["priority"] == priority]
        if not section_items:
            continue
        lines.append(f"### {priority_heading(priority)}")
        lines.append("")
        for item in section_items:
            why = next(
                (
                    finding["why"]
                    for site_findings in findings_by_site.values()
                    for finding in site_findings
                    if finding["key"] == item["key"]
                ),
                "",
            )
            lines.append(f"#### {item['title']}")
            lines.append("")
            lines.append(f"Priority: {priority_label(item['priority'])}")
            lines.append("")
            append_action_table(
                lines,
                item["key"],
                item["priority"],
                item["team"],
                item["actions"],
                why,
                item["actions"][0] if item["actions"] else item["title"],
                format_list(item["sites"]),
                None,
                item["optional"],
            )
        lines.append("")
    if include_custom_mcp_note:
        lines.append("### Strategic Build Included In This Plan: Custom MCP")
        lines.append("")
        append_action_table(
            lines,
            "custom_mcp",
            "P2",
            "Product + Dev",
            [
                "Keep Shopify's native MCP for catalog search, cart operations, product lookup, and policy answers.",
                "Use a separate custom MCP only for logic Shopify's native MCP does not cover well enough, such as recommendation explanations, product comparison, sizing or fit interpretation, and brand guidance.",
                "Do not rebuild standard catalog or cart tools inside the custom MCP if Shopify's native MCP already handles them.",
                "Priority note: Optional after the core fixes are stable.",
            ],
            "Shopify's native Storefront MCP is useful for baseline commerce tasks, but Shopify does not let you edit that native MCP to add brand-specific advisory logic.",
            "If AI-assisted shopping is a priority, plan a separate custom MCP for recommendation reasons, comparison logic, sizing or fit advice, brand facts, and market-specific guidance.",
            label_text,
            None,
            True,
        )
    lines.append("## Site-by-Site Action Plan")
    lines.append("")
    for domain, site_findings in findings_by_site.items():
        site_label = site_findings[0]["site_label"] if site_findings else domain
        lines.append(f"### {site_label}")
        lines.append("")
        for finding in site_findings:
            lines.append(f"#### {finding['title']}")
            lines.append("")
            lines.append(f"Priority: {priority_label(finding['priority'])}")
            lines.append("")
            append_action_table(
                lines,
                finding["key"],
                finding["priority"],
                finding["team"],
                finding["actions"],
                finding["why"],
                finding["recommendation"],
                None,
                finding["evidence"],
                finding["optional"],
            )
        lines.append("")
    lines.append("## Conclusion")
    lines.append("")
    lines.append(f"- Most of the important AI-readiness foundation work applies across all {total_sites} brands, not just one site.")
    lines.append("- The biggest shared themes are AI product citation coverage, homepage schema depth, homepage title/meta/H1 consistency, deeper product and collection schema, schema-to-page alignment, trust/help discoverability, and AI brand visibility.")
    lines.append(f"- The implementation scope is not identical across the {total_sites} stores, so the rollout should use one shared foundation roadmap plus brand-specific fixes.")
    lines.append("- The per-site action tables should drive sequencing: apply shared P0 fixes first, then close larger store-specific PDP and schema gaps.")
    if any("aari" in label.lower() for label in labels):
        lines.append("- Aari Clothing is included as an incremental workstream in the same roadmap and should be tracked with the same QA and re-test cadence.")
    lines.append("")
    lines.append("## Delivery Phases")
    lines.append("")
    if include_custom_mcp_note:
        lines.append(f"Current total delivery estimate is {total_effort_label} including custom MCP, plus planning, reviews, approvals, content production, and iteration overhead.")
    else:
        lines.append(f"Current core implementation estimate is {core_effort}, plus planning, reviews, approvals, content production, and iteration overhead.")
    lines.append("Visibility re-testing should run every 2 weeks after JSON-LD or other structured-data updates so the team can measure whether schema changes are improving brand recall and product citation outcomes.")
    lines.append("")
    lines.append('<table>')
    lines.append('<colgroup>')
    lines.append('<col style="width: 18%;">')
    lines.append('<col style="width: 42%;">')
    lines.append('<col style="width: 40%;">')
    lines.append('</colgroup>')
    lines.append('<thead>')
    lines.append('<tr>')
    lines.append('<th style="vertical-align: top; text-align: left;">Phase</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Focus</th>')
    lines.append('<th style="vertical-align: top; text-align: left;">Expected outcome</th>')
    lines.append('</tr>')
    lines.append('</thead>')
    lines.append('<tbody>')
    lines.append('<tr>')
    lines.append('<td style="vertical-align: top;">Phase 1</td>')
    lines.append('<td style="vertical-align: top;">Discovery, backlog shaping, implementation planning, and the first pass of homepage metadata, homepage entity/schema, and core schema/page-alignment fixes.</td>')
    lines.append('<td style="vertical-align: top;">The program starts with an agreed roadmap and the most important brand and site signals begin to stabilize.</td>')
    lines.append('</tr>')
    lines.append('<tr>')
    lines.append('<td style="vertical-align: top;">Phase 2</td>')
    lines.append('<td style="vertical-align: top;">PDP schema expansion, product metadata cleanup, breadcrumb improvements, stronger trust, help, shipping, and policy discoverability, plus the first biweekly visibility re-test after structured-data changes.</td>')
    lines.append('<td style="vertical-align: top;">Product pages become cleaner machine-readable assets and easier for AI systems to interpret accurately, with early visibility feedback starting to validate the direction of change.</td>')
    lines.append('</tr>')
    lines.append('<tr>')
    lines.append('<td style="vertical-align: top;">Phase 3</td>')
    lines.append('<td style="vertical-align: top;">Richer PDP content rollout, apparel attribute depth, related-product logic, follow-up-answer content, secondary feed or structured-output improvements, and continued biweekly visibility re-tests after schema/content releases.</td>')
    lines.append('<td style="vertical-align: top;">The storefronts become more citation-worthy and more capable of answering real shopper follow-up questions, while visibility tracking shows whether those updates are converting into stronger recall and citations.</td>')
    lines.append('</tr>')
    lines.append('<tr>')
    lines.append('<td style="vertical-align: top;">Phase 4</td>')
    lines.append('<td style="vertical-align: top;">Brand-specific gap closure for the lower-scoring stores, plus refinement of remaining schema, metadata, and content gaps that were deprioritized earlier, with biweekly visibility re-tests continuing after major JSON-LD updates and custom MCP discovery or architecture work starting.</td>')
    lines.append('<td style="vertical-align: top;">Lower-scoring stores close their larger readiness gaps while stronger stores get refinement rather than rework, and the custom MCP workstream starts inside the same overall program.</td>')
    lines.append('</tr>')
    lines.append('<tr>')
    lines.append('<td style="vertical-align: top;">Phase 5</td>')
    lines.append('<td style="vertical-align: top;">Biweekly visibility re-testing, storefront re-audit, implementation QA, stakeholder review of what improved, and the main custom MCP design and build phase.</td>')
    lines.append('<td style="vertical-align: top;">The team can verify whether citation coverage, brand recall, and storefront readiness materially improved, while the custom MCP moves from requirements into active implementation.</td>')
    lines.append('</tr>')
    lines.append('<tr>')
    lines.append('<td style="vertical-align: top;">Phase 6</td>')
    lines.append('<td style="vertical-align: top;">Custom MCP build completion, QA, pilot rollout, final optimisation cycle, content polish, documentation, repeatable QA handoff, and prioritisation of any follow-on items.</td>')
    lines.append('<td style="vertical-align: top;">The engagement closes with storefront-readiness improvements and a custom MCP workstream delivered inside the same delivery program.</td>')
    lines.append('</tr>')
    lines.append('</tbody>')
    lines.append('</table>')
    lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--storefront-audit", default=str(DEFAULT_STOREFRONT_AUDIT))
    parser.add_argument("--admin-audit", default=str(DEFAULT_ADMIN_AUDIT))
    parser.add_argument("--visibility-report", default="")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH))
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    storefront_path = Path(args.storefront_audit)
    admin_path = Path(args.admin_audit)
    visibility_path = Path(args.visibility_report) if args.visibility_report else None
    config_path = Path(args.config)

    storefront_payload = load_json(storefront_path)
    admin_payload = load_json(admin_path) if admin_path.exists() else None
    visibility_payload = load_visibility_workbook(visibility_path) if visibility_path and visibility_path.exists() else None
    config = load_config(config_path)
    labels = label_map(config, storefront_payload.get("results", []))
    generator_module = load_existing_module(Path("scripts/generate_shopify_audit_workbook.py"))

    findings_by_site = build_site_findings(storefront_payload, admin_payload, generator_module, labels)
    visibility_findings = build_visibility_findings(visibility_payload, config, labels)
    for domain, items in visibility_findings.items():
        findings_by_site[domain] = sorted(findings_by_site.get(domain, []) + items, key=severity_key)
    cross_site = build_cross_site_summary(findings_by_site)
    consolidated_actions = generator_module.build_consolidated_actions(storefront_payload.get("results", []), [
        {check["key"]: generator_module.eval_check(check["key"], generator_module.build_metrics(site))
         for check in generator_module.DETAIL_CHECKS}
        for site in storefront_payload.get("results", [])
    ])
    effort_summary = generator_module.consolidated_action_effort_summary(consolidated_actions)
    include_custom_mcp_note = True
    generated_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"
    markdown = render_markdown(
        findings_by_site,
        cross_site,
        consolidated_actions,
        effort_summary,
        generated_at,
        str(visibility_path) if visibility_path and visibility_path.exists() else None,
        include_custom_mcp_note,
    )

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(markdown, encoding="utf-8")


if __name__ == "__main__":
    main()
