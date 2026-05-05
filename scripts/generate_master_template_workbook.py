#!/usr/bin/env python3

from __future__ import annotations

import argparse
import importlib.util
import json
from collections import Counter, defaultdict
from copy import copy
from datetime import datetime
from pathlib import Path
from statistics import mean
from typing import Any, Dict, List, Tuple

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment

DEFAULT_CONFIG_PATH = Path("configs/audit_config.json")

SITE_ORDER = ["trailberg.com", "lorenzo.world", "dreamisfree.com"]
SITE_LABELS = {
    "trailberg.com": "Trailberg",
    "lorenzo.world": "Lorenzo",
    "dreamisfree.com": "Dream Is Free",
}
VISIBILITY_SHEET_MAP = {
    "trailberg.com": "Trailberg_Visibility Tracker",
    "lorenzo.world": "Lorenzo_Visibility Tracker",
    "dreamisfree.com": "Dream Is Free_Visibility Tracke",
}
VISIBILITY_REPORT_MAP = {
    "trailberg.com": "Trailberg",
    "lorenzo.world": "Lorenzo",
    "dreamisfree.com": "Dreamisfree",
}
BRAND_ALIASES = {
    "trailberg.com": ["trailberg"],
    "lorenzo.world": ["lorenzo"],
    "dreamisfree.com": ["dreamisfree", "dream is free"],
}
PRIVATE_CHECK_KEYS = {
    "admin_only",
    "out_of_scope",
    "option_consistency",
    "metafield_namespaces",
    "metafield_quality",
    "metaobject_usage",
    "taxonomy_assignment",
    "collection_logic",
    "related_product_data",
    "inventory_tracking",
    "inventory_policy",
    "inventory_locations",
    "theme_architecture",
    "webhook_setup",
    "comparison_readiness",
    "manual_product_lifecycle",
    "manual_inventory_freshness",
    "manual_inventory_experience",
    "manual_commerce_logic",
    "manual_search_analytics",
    "manual_content_governance",
    "manual_agent_analytics",
    "manual_ai_governance",
    "manual_ai_quality",
    "manual_ai_tooling",
    "manual_ai_benchmarks",
    "manual_data_ownership",
    "manual_ai_ecosystem",
}


def load_config(config_path: str | None) -> None:
    global SITE_ORDER, SITE_LABELS, VISIBILITY_SHEET_MAP, VISIBILITY_REPORT_MAP, BRAND_ALIASES
    path = Path(config_path or DEFAULT_CONFIG_PATH)
    if not path.exists():
        return
    payload = json.loads(path.read_text(encoding="utf-8"))
    sites = payload.get("sites", [])
    if not sites:
        return
    SITE_ORDER = [site["domain"] for site in sites]
    SITE_LABELS = {site["domain"]: site["label"] for site in sites}
    VISIBILITY_SHEET_MAP = {site["domain"]: site["visibility_sheet"] for site in sites}
    VISIBILITY_REPORT_MAP = {site["domain"]: site["visibility_report_sheet"] for site in sites}
    BRAND_ALIASES = {
        site["domain"]: site.get("brand_aliases", [site["label"].lower()])
        for site in sites
    }


def load_existing_module(path: str):
    spec = importlib.util.spec_from_file_location("storefront_gen", path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def clone_style(source, target) -> None:
    target._style = copy(source._style)
    target.number_format = source.number_format
    target.alignment = copy(source.alignment)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.protection = copy(source.protection)


def copy_sheet_exact(source_ws, target_ws) -> None:
    clear_range(target_ws, 1, max(target_ws.max_row, source_ws.max_row) + 5, 1, max(target_ws.max_column, source_ws.max_column) + 5)
    for row in source_ws.iter_rows():
        for source_cell in row:
            target_cell = target_ws.cell(source_cell.row, source_cell.column, source_cell.value)
            clone_style(source_cell, target_cell)
            if source_cell.has_style:
                target_cell.number_format = source_cell.number_format
            target_cell.alignment = copy(source_cell.alignment)

    target_ws.freeze_panes = source_ws.freeze_panes
    target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines

    for key, dim in source_ws.column_dimensions.items():
        target_dim = target_ws.column_dimensions[key]
        target_dim.width = dim.width
        target_dim.hidden = dim.hidden
        target_dim.bestFit = dim.bestFit

    for idx, dim in source_ws.row_dimensions.items():
        target_dim = target_ws.row_dimensions[idx]
        target_dim.height = dim.height
        target_dim.hidden = dim.hidden

    for merged in list(target_ws.merged_cells.ranges):
        target_ws.unmerge_cells(str(merged))
    for merged in source_ws.merged_cells.ranges:
        target_ws.merge_cells(str(merged))


def clear_range(ws, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            cell.value = None


def is_generic_query(row: Dict[str, Any], target_brand: str) -> bool:
    intent = str(row.get("intent") or "").strip().lower()
    query = str(row.get("query") or "").strip().lower()
    target = target_brand.strip().lower()
    aliases = {target}
    compact = "".join(ch for ch in target if ch.isalnum())
    if compact:
        aliases.add(compact)
    website = str(row.get("target_brand_website") or "").strip().lower()
    for domain, configured_aliases in BRAND_ALIASES.items():
        if website.endswith(domain) or target in {alias.lower() for alias in configured_aliases}:
            aliases.update(alias.lower() for alias in configured_aliases)
    if intent == "branded":
        return False
    if any(alias and alias in query for alias in aliases):
        return False
    return True


def compute_visibility_metrics(rows: List[Dict[str, Any]], target_brand: str) -> Dict[str, Dict[str, Dict[str, Any]]]:
    target_brand_lower = target_brand.strip().lower()
    generic_rows = [
        row
        for row in rows
        if is_generic_query(row, target_brand) and str(row.get("status") or "").strip().lower() != "error"
    ]
    platforms = defaultdict(list)
    for row in generic_rows:
        platforms[row["platform"]].append(row)
    platforms["All Platforms"] = generic_rows[:]

    def visibility_score(mention_rate: float, top3_rate: float, avg_rank: float | None) -> float:
        rank_component = 0.0 if avg_rank is None else max(0.0, (6.0 - avg_rank) / 5.0) * 20.0
        return round((mention_rate * 0.5) + (top3_rate * 0.3) + rank_component, 2)

    results: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for platform, platform_rows in platforms.items():
        total = len(platform_rows) or 1
        brand_data: Dict[str, Dict[str, Any]] = defaultdict(lambda: {"positions": [], "appearances": 0})
        for row in platform_rows:
            mention = str(row["brand_mentioned"]).strip().upper() == "Y"
            if mention:
                brand_data[target_brand]["appearances"] += 1
                if row["brand_position"] is not None:
                    brand_data[target_brand]["positions"].append(float(row["brand_position"]))
            brands = [item.strip() for item in str(row["top3"] or "").split(",") if item.strip()]
            for idx, brand in enumerate(brands, start=1):
                brand_data[brand]["appearances"] += 1
                brand_data[brand]["positions"].append(float(idx))

        target_entry = {
            "brand": target_brand,
            "mention_rate": 0.0,
            "avg_rank": None,
            "top3_rate": 0.0,
            "visibility_score": 0.0,
            "sample_size": len(platform_rows),
        }
        if target_brand in brand_data:
            positions = brand_data[target_brand]["positions"]
            target_entry = {
                "brand": target_brand,
                "mention_rate": round((brand_data[target_brand]["appearances"] / total) * 100, 2),
                "avg_rank": round(mean(positions), 2) if positions else None,
                "top3_rate": round((sum(1 for pos in positions if pos <= 3) / total) * 100, 2) if positions else 0.0,
                "visibility_score": visibility_score(
                    round((brand_data[target_brand]["appearances"] / total) * 100, 2),
                    round((sum(1 for pos in positions if pos <= 3) / total) * 100, 2) if positions else 0.0,
                    round(mean(positions), 2) if positions else None,
                ),
                "sample_size": len(platform_rows),
            }

        competitors = []
        for brand, data in brand_data.items():
            if brand.strip().lower() == target_brand_lower:
                continue
            positions = data["positions"]
            mention_rate = round((data["appearances"] / total) * 100, 2)
            avg_rank = round(mean(positions), 2) if positions else None
            top3_rate = round((sum(1 for pos in positions if pos <= 3) / total) * 100, 2) if positions else 0.0
            competitors.append(
                {
                    "brand": brand,
                    "mention_rate": mention_rate,
                    "avg_rank": avg_rank,
                    "top3_rate": top3_rate,
                    "visibility_score": visibility_score(mention_rate, top3_rate, avg_rank),
                    "sample_size": len(platform_rows),
                }
            )
        competitors.sort(key=lambda item: (-item["visibility_score"], item["brand"].lower()))
        results[platform] = {
            "target": target_entry,
            "top_brands": [target_entry] + competitors[:4],
        }
    return results


def load_visibility_rows(visibility_wb) -> Dict[str, List[Dict[str, Any]]]:
    out: Dict[str, List[Dict[str, Any]]] = {}
    for domain, sheet_name in VISIBILITY_REPORT_MAP.items():
        ws = visibility_wb[sheet_name]
        header_map = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header is None or str(header).strip() == "":
                continue
            header_map[str(header).strip().lower()] = col

        def cell_value(row_idx: int, *keys: str):
            for key in keys:
                col = header_map.get(key.lower())
                if col:
                    return ws.cell(row_idx, col).value
            return None

        rows = []
        row_idx = 2
        while True:
            values = [ws.cell(row_idx, col).value for col in range(1, max(ws.max_column, 11) + 1)]
            if not any(v is not None and v != "" for v in values):
                break
            rows.append(
                {
                    "checked_on": cell_value(row_idx, "checked on", "date"),
                    "platform": cell_value(row_idx, "platform"),
                    "model_used": cell_value(row_idx, "model used"),
                    "query": cell_value(row_idx, "query"),
                    "intent": cell_value(row_idx, "intent"),
                    "target_brand_name": cell_value(row_idx, "target brand name"),
                    "target_brand_website": cell_value(row_idx, "target brand website"),
                    "brand_mentioned": cell_value(row_idx, "brand mentioned"),
                    "brand_position": cell_value(row_idx, "brand position"),
                    "top3": cell_value(row_idx, "top 3 brands in order"),
                    "sentiment": cell_value(row_idx, "sentiment"),
                    "response_type": cell_value(row_idx, "response type"),
                    "notes": cell_value(row_idx, "notes"),
                    "citations_json": cell_value(row_idx, "citations_json"),
                    "status": cell_value(row_idx, "status"),
                    "error": cell_value(row_idx, "error"),
                }
            )
            row_idx += 1
        out[domain] = rows
    return out


def normalize_domain_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.removeprefix("https://").removeprefix("http://")
    return text.rstrip("/")


def parse_citations(value: Any) -> List[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if value in (None, ""):
        return []
    try:
        parsed = json.loads(str(value))
        if isinstance(parsed, list):
            return [str(item).strip() for item in parsed if str(item).strip()]
    except Exception:
        pass
    return []


def visibility_remark(prefix: str, detail: str) -> str:
    return f"{prefix}; {detail}" if detail else prefix


def summarize_visibility_insights(rows: List[Dict[str, Any]], target_brand: str, target_domain: str) -> Dict[str, Dict[str, Any]]:
    completed_rows = [row for row in rows if str(row.get("status") or "").strip().lower() != "error"]
    error_rows = [row for row in rows if str(row.get("status") or "").strip().lower() == "error"]
    metrics = compute_visibility_metrics(completed_rows, target_brand)
    target_stats = metrics.get("All Platforms", {}).get(
        "target",
        {"mention_rate": 0.0, "avg_rank": None, "top3_rate": 0.0, "visibility_score": 0.0},
    )
    competitors = [
        item
        for item in metrics.get("All Platforms", {}).get("top_brands", [])
        if str(item.get("brand") or "").strip().lower() != target_brand.strip().lower()
    ]
    top_competitor = competitors[0] if competitors else None

    normalized_domain = normalize_domain_text(target_domain)
    citation_rows = 0
    product_citation_rows = 0
    guide_citation_rows = 0
    sentiments = {"positive": 0, "neutral": 0, "negative": 0, "not_mentioned": 0}
    notes = []
    for row in completed_rows:
        citations = parse_citations(row.get("citations_json"))
        target_citations = [item for item in citations if normalized_domain and normalized_domain in item.lower()]
        if target_citations:
            citation_rows += 1
        if any("/products/" in item.lower() for item in target_citations):
            product_citation_rows += 1
        if any("/pages/" in item.lower() or "/blogs/" in item.lower() for item in target_citations):
            guide_citation_rows += 1
        sentiment = str(row.get("sentiment") or "not_mentioned").strip().lower() or "not_mentioned"
        sentiments[sentiment if sentiment in sentiments else "not_mentioned"] += 1
        note = str(row.get("notes") or "").strip()
        if note:
            notes.append(note)

    sample_size = len(completed_rows)
    citation_rate = round((citation_rows / sample_size) * 100, 2) if sample_size else 0.0
    product_citation_rate = round((product_citation_rows / sample_size) * 100, 2) if sample_size else 0.0
    guide_citation_rate = round((guide_citation_rows / sample_size) * 100, 2) if sample_size else 0.0
    dominant_sentiment = max(sentiments, key=sentiments.get) if sample_size else "not_mentioned"
    representative_note = next((note for note in notes if note), "")

    brand_detail = f"top competitor {top_competitor['brand']}" if top_competitor else "no strong competitor pattern yet"
    if target_stats["mention_rate"] >= 60:
        brand_summary = visibility_remark("Strong brand visibility", brand_detail)
    elif target_stats["mention_rate"] >= 30:
        brand_summary = visibility_remark("Mixed brand visibility", brand_detail)
    else:
        brand_summary = visibility_remark("Weak brand visibility", brand_detail)

    if product_citation_rate >= 25:
        product_summary = "Products are surfacing in AI-visible citations."
    elif product_citation_rate > 0:
        product_summary = "Some product visibility exists, but PDP citation coverage is still thin."
    else:
        product_summary = "No target product citations were detected in the current AI runs."

    if citation_rate >= 40:
        citation_summary = "AI platforms are citing the site regularly."
    elif citation_rate > 0:
        citation_summary = "AI platforms cite the site occasionally, but source authority is still inconsistent."
    else:
        citation_summary = "No target-domain citations were captured."

    if dominant_sentiment == "positive":
        framing_summary = "Brand framing is generally positive."
    elif dominant_sentiment == "neutral":
        framing_summary = "Brand framing is mostly neutral."
    elif dominant_sentiment == "negative":
        framing_summary = "Brand framing is skewing negative."
    else:
        framing_summary = "The brand is not being framed consistently because it is often not mentioned."

    competitor_summary = (
        f"{top_competitor['brand']} appears most often alongside the brand."
        if top_competitor
        else "No recurring competitor exposure pattern was detected."
    )

    return {
        "brand": {
            "mention_rate": target_stats["mention_rate"],
            "avg_rank": target_stats["avg_rank"],
            "top3_rate": target_stats["top3_rate"],
            "visibility_score": target_stats["visibility_score"],
            "sample_size": sample_size,
            "errors": len(error_rows),
            "remark": brand_summary,
            "note": representative_note,
        },
        "product": {
            "product_citation_rate": product_citation_rate,
            "sample_size": sample_size,
            "errors": len(error_rows),
            "remark": product_summary,
            "note": representative_note,
        },
        "citation": {
            "citation_rate": citation_rate,
            "product_citation_rate": product_citation_rate,
            "guide_citation_rate": guide_citation_rate,
            "sample_size": sample_size,
            "errors": len(error_rows),
            "remark": citation_summary,
            "note": representative_note,
        },
        "framing": {
            "dominant_sentiment": dominant_sentiment,
            "positive": sentiments["positive"],
            "neutral": sentiments["neutral"],
            "negative": sentiments["negative"],
            "sample_size": sample_size,
            "errors": len(error_rows),
            "remark": framing_summary,
            "note": representative_note,
        },
        "competitors": {
            "top_competitor": top_competitor["brand"] if top_competitor else None,
            "top_competitor_mention_rate": top_competitor["mention_rate"] if top_competitor else 0.0,
            "sample_size": sample_size,
            "errors": len(error_rows),
            "remark": competitor_summary,
            "note": representative_note,
        },
    }


def rating(score: float) -> str:
    if score >= 4.5:
        return "Strong"
    if score >= 3.0:
        return "Mixed"
    if score > 0:
        return "Weak"
    return "Out of scope"


def compact_site_summary(result: Dict[str, Any]) -> str:
    score = result["score"]
    status = result["status"]
    if status == "OUT OF SCOPE":
        return "OUT OF SCOPE"
    if status == "MANUAL":
        return "MANUAL"
    return f"{status} ({score})"


def average_score(items: List[Dict[str, Any]]) -> float:
    scores = [item["score"] for item in items if item["status"] not in {"MANUAL", "OUT OF SCOPE"}]
    return round(sum(scores) / len(scores), 2) if scores else 0.0


def summary_text(items: List[Dict[str, Any]]) -> str:
    non_manual = [item for item in items if item["status"] not in {"MANUAL", "OUT OF SCOPE"}]
    if not non_manual and any(item["status"] == "OUT OF SCOPE" for item in items):
        return "Out of scope for a public-render audit."
    if not non_manual and any(item["status"] == "MANUAL" for item in items):
        return "Requires additional operational evidence beyond public storefront signals."
    if not non_manual:
        return "No public-render evidence was available for this check."
    if all(item["status"] == "PASS" for item in non_manual):
        return "Consistently strong across the 3 public storefronts."
    if any(item["status"] == "FAIL" for item in non_manual):
        return "Mixed public signals; at least one storefront needs direct remediation."
    return "Public signals are present but inconsistent across the 3 storefronts."


def recommended_effort(items: List[Dict[str, Any]]) -> str:
    avg_score = average_score(items)
    if not [item for item in items if item["status"] not in {"MANUAL", "OUT OF SCOPE"}]:
        if any(item["status"] == "OUT OF SCOPE" for item in items):
            return "Out of scope"
        return "Manual review"
    if avg_score >= 4.5:
        return "-"
    if avg_score >= 3:
        return "1-2"
    if avg_score > 0:
        return "2-4"
    return "Out of scope"


def business_impact(check_key: str) -> str:
    if check_key in {"homepage_status", "robots", "sitemap", "product_metadata", "product_schema_presence", "products_json"}:
        return "High"
    if check_key in {
        "homepage_metadata",
        "collection_metadata",
        "machine_feeds",
        "ai_bot_policy",
        "llms_txt",
        "manual_agent_analytics",
        "manual_ai_governance",
        "manual_ai_quality",
        "manual_ai_benchmarks",
        "comparison_readiness",
        "inventory_tracking",
        "inventory_policy",
        "inventory_locations",
        "manual_inventory_freshness",
        "manual_inventory_experience",
        "manual_data_ownership",
    }:
        return "Medium"
    if check_key in {"admin_only", "out_of_scope"}:
        return "N/A"
    return "Medium"


def target_outcome(check_key: str) -> str:
    mapping = {
        "homepage_status": "Stable, crawlable homepage",
        "homepage_metadata": "Clear homepage entity signals",
        "homepage_schema": "Homepage Organization/WebSite schema",
        "homepage_schema_depth": "Intent-rich homepage entity schema",
        "robots": "Intentional crawler directives",
        "sitemap": "Complete public URL discovery",
        "url_structure": "Clean entity-separated URL patterns",
        "collections_exist": "Discoverable collection landing pages",
        "collection_metadata": "Category pages with useful metadata",
        "collection_content": "Category pages with semantic context",
        "product_metadata": "PDPs eligible for citation",
        "product_schema_presence": "Product schema on every PDP",
        "product_schema_completeness": "Commercial fields inside schema",
        "schema_depth": "Deep product and collection schema",
        "breadcrumbs": "Clear product hierarchy signals",
        "image_alt": "Better accessibility + machine cues",
        "products_json": "Public machine-readable feed",
        "semantic_depth": "PDPs that answer follow-up questions",
        "search_facets": "Visible search and filter semantics",
        "trust_content": "Crawlable trust/support content",
        "follow_up_readiness": "Public facts sufficient for AI answers",
        "related_products": "Visible alternatives and cross-sell logic",
        "clothing_attribute_depth": "Fabric/fit/weather/style field depth",
        "llms_txt": "Explicit AI-consumption policy",
        "ai_bot_policy": "No accidental AI crawler blocks",
        "machine_feeds": "Multiple structured ingestion channels",
        "schema_alignment": "Visible and machine-readable facts aligned",
        "regional_signals": "Correct market/locale context",
        "mcp": "Native Shopify storefront MCP availability",
        "mcp_depth": "Usable Shopify MCP tool contract and RPC transport",
        "custom_mcp": "Brand-specific custom MCP opportunity",
        "option_consistency": "Consistent variant option vocabularies",
        "metafield_quality": "Publicly rendered product facts",
        "metafield_namespaces": "Consistent public attribute naming",
        "metaobject_usage": "Reusable rendered semantic content blocks",
        "taxonomy_assignment": "Clear rendered category semantics",
        "collection_logic": "Explainable collection membership",
        "related_product_data": "Explicit related and upsell data",
        "inventory_tracking": "Variant-level inventory truth",
        "inventory_policy": "Intentional oversell/backorder policies",
        "inventory_locations": "Clear multi-location inventory model",
        "theme_architecture": "Documented theme and app boundaries",
        "webhook_setup": "Observable data-change integrations",
        "comparison_readiness": "Structured product comparison data",
        "manual_product_lifecycle": "Clear lifecycle and publication rules",
        "manual_inventory_freshness": "Documented stock freshness and ownership",
        "manual_inventory_experience": "Consistent stock messaging and availability answers",
        "manual_commerce_logic": "Predictable pricing, shipping, and tax rules",
        "manual_search_analytics": "Search behaviour logging and improvement loop",
        "manual_content_governance": "Content review and retirement process",
        "manual_agent_analytics": "AI journey and failure observability",
        "manual_ai_governance": "Named AI discovery ownership and review process",
        "manual_ai_quality": "Repeatable AI quality checks and rollback path",
        "manual_ai_tooling": "Tooling, experimentation, and cost visibility",
        "manual_ai_benchmarks": "Benchmarks and before/after tracking",
        "manual_data_ownership": "Clear ownership and traceability for product truth",
        "manual_ai_ecosystem": "Externally consistent brand and usage governance",
        "admin_only": "Hidden platform configuration out of scope",
        "out_of_scope": "Not visible in public frontend render",
    }
    return mapping.get(check_key, "Public storefront readiness")


def vector_db_flag(check_key: str) -> str:
    if check_key in {"product_schema_presence", "product_schema_completeness", "products_json", "machine_feeds", "semantic_depth", "trust_content", "metafield_quality", "metaobject_usage", "comparison_readiness"}:
        return "Yes"
    if check_key in {"admin_only", "out_of_scope"}:
        return "N/A"
    return "Indirect"


def manual_review_result(
    generator_module,
    evidence: str,
    recommendation: str,
    observability: str = "Manual review",
    sources: str = "Manual audit",
) -> Dict[str, Any]:
    return generator_module.result("MANUAL", 0, evidence, recommendation, observability=observability, sources=sources)


def manual_review_checks(generator_module) -> Dict[str, Dict[str, Any]]:
    return {
        "option_consistency": manual_review_result(generator_module, "Requires variant-option review in Shopify Admin or product exports.", "Verify that size, colour, material, and other option vocabularies use one consistent format across the catalogue.", sources="Admin option audit"),
        "metafield_namespaces": manual_review_result(generator_module, "Requires Admin metafield namespace review.", "Check metafield namespaces for clear naming, consistent usage, and low theme-hack leakage.", sources="Admin metafield audit"),
        "metafield_quality": manual_review_result(generator_module, "Requires Admin metafield definitions and sampled product records.", "Verify that important product facts are stored in structured fields, not only implied in copy or hidden theme settings.", sources="Admin metafield audit"),
        "metaobject_usage": manual_review_result(generator_module, "Requires GraphQL metaobject definition review.", "Check whether repeatable semantic content such as sizing, care, materials, and FAQs is modeled with reusable metaobjects.", sources="Admin GraphQL audit"),
        "taxonomy_assignment": manual_review_result(generator_module, "Requires Shopify product category review in Admin.", "Confirm that active products use the correct Shopify taxonomy and that assignment rules are documented.", sources="Admin taxonomy audit"),
        "collection_logic": manual_review_result(generator_module, "Requires collection-rule or manual-curation review in Admin.", "Document why products belong to each collection and whether membership is rule-based, manual, or mixed.", sources="Admin collection audit"),
        "related_product_data": manual_review_result(generator_module, "Requires Search & Discovery / merchandising rule review.", "Verify that related, upsell, and alternative recommendations are data-backed and explainable.", sources="Merchandising review"),
        "inventory_tracking": manual_review_result(generator_module, "Requires variant inventory settings and stock-system ownership review.", "Confirm that inventory is tracked at variant level and that Shopify's stock source-of-truth model is explicit.", sources="Inventory operations review"),
        "inventory_policy": manual_review_result(generator_module, "Requires inventory policy review in Admin and operational docs.", "Review continue-selling, oversell prevention, backorder, and pre-order rules for consistency and customer clarity.", sources="Inventory policy review"),
        "inventory_locations": manual_review_result(generator_module, "Requires multi-location inventory and fulfilment rule review.", "Check how stock behaves across locations, warehouses, regions, and fulfilment methods.", sources="Location inventory review"),
        "theme_architecture": manual_review_result(generator_module, "Requires theme/app architecture and ownership review.", "Document what logic lives in theme code, apps, Shopify Functions, and any headless surfaces.", sources="Architecture review"),
        "webhook_setup": manual_review_result(generator_module, "Requires webhook and integration reliability review.", "Verify that change events are observable, idempotent, and traceable back to their source systems.", sources="Integration review"),
        "comparison_readiness": manual_review_result(generator_module, "Requires structured attribute and recommendation-data review.", "Check whether products can be compared, recommended, and caveated using explicit data rather than merch rules alone.", sources="Product reasoning review"),
        "manual_product_lifecycle": manual_review_result(generator_module, "Requires merchandising and catalogue-governance review.", "Define lifecycle states such as new, core, seasonal, discontinued, plus clear rules for draft, active, and archived products.", sources="Catalogue governance review"),
        "manual_inventory_freshness": manual_review_result(generator_module, "Requires stock-sync and operational documentation review.", "Document stock freshness, sync delays, source-of-truth ownership, and how availability confidence is determined.", sources="Stock operations review"),
        "manual_inventory_experience": manual_review_result(generator_module, "Requires journey QA across PDP, cart, and checkout.", "Verify that stock messaging, fulfilment expectations, and 'available right now' answers stay consistent across the customer journey.", sources="Journey QA review"),
        "manual_commerce_logic": manual_review_result(generator_module, "Requires pricing, shipping, tax, and function-rule review.", "Document discount, shipping, market, and tax logic in plain English so outcomes are predictable and testable.", sources="Commerce operations review"),
        "manual_search_analytics": manual_review_result(generator_module, "Requires search analytics and reporting review.", "Check whether poor searches, zero-result terms, and AI-style natural-language searches are logged and improved over time.", sources="Search analytics review"),
        "manual_content_governance": manual_review_result(generator_module, "Requires editorial governance review.", "Document how content is reviewed, refreshed, retired, and connected to product truth over time.", sources="Content governance review"),
        "manual_agent_analytics": manual_review_result(generator_module, "Requires analytics instrumentation and reporting review.", "Manually review whether AI-agent traffic, success/failure signals, uncertainty, and AI-led journeys are visible in analytics.", sources="Agent analytics review"),
        "manual_ai_governance": manual_review_result(generator_module, "Requires owner, workflow, and review-process confirmation.", "Confirm shared responsibility, named ownership, and a clear root-cause process for AI discovery failures.", sources="AI governance review"),
        "manual_ai_quality": manual_review_result(generator_module, "Requires repeatable evaluation and release-process review.", "Define recurring test prompts, drift checks, refresh plans, regression tests, and rollback procedures for AI behaviour.", sources="AI quality review"),
        "manual_ai_tooling": manual_review_result(generator_module, "Requires tooling and cost review.", "Document what tools are used to inspect AI behaviour, how experiments are run safely, and how AI tooling costs are tracked.", sources="Tooling review"),
        "manual_ai_benchmarks": manual_review_result(generator_module, "Requires benchmark and KPI review.", "Track baseline, current state, success/confidence/failure rates, and stakeholder-facing progress over time.", sources="Benchmark review"),
        "manual_data_ownership": manual_review_result(generator_module, "Requires systems and ownership mapping.", "Clarify which system owns product truth, where derived data originates, and how changes are audited or rolled back.", sources="Data ownership review"),
        "manual_ai_ecosystem": manual_review_result(generator_module, "Requires partner-feed, rights, and external-consistency review.", "Review brand consistency across AI assistants, partner/platform feeds, usage rights, regional context, and the trusted source for brand facts.", sources="AI ecosystem review"),
    }


def classify_readiness_row(domain: str, what_check: str) -> str:
    text = f"{domain} {what_check}".lower()
    if "brand visibility" in text:
        return "visibility_brand"
    if "product visibility" in text:
        return "visibility_product"
    if "citation" in text:
        return "visibility_citation"
    if "brand framing" in text:
        return "visibility_framing"
    if "competitive exposure" in text:
        return "visibility_competitors"
    if "canonical product structure" in text:
        return "url_structure"
    if "variant logic" in text:
        return "schema_depth"
    if "option consistency" in text:
        return "option_consistency"
    if "product metafields" in text:
        return "metafield_quality"
    if "metaobjects" in text:
        return "metaobject_usage"
    if "shopify product category" in text:
        return "taxonomy_assignment"
    if "collection logic" in text:
        return "collection_logic"
    if "collections" in text:
        return "collection_content"
    if "related products" in text:
        return "related_product_data"
    if "product lifecycle" in text:
        return "manual_product_lifecycle"
    if "inventory accuracy" in text:
        return "inventory_tracking"
    if "multi-location" in text:
        return "inventory_locations"
    if "backorder" in text or "pre-order" in text:
        return "inventory_policy"
    if "stock messaging" in text:
        return "manual_inventory_experience"
    if "discount" in text or "shipping" in text or "tax" in text:
        return "manual_commerce_logic"
    if "product comparability" in text or "suitability signals" in text or "limitations" in text or "use cases" in text:
        return "comparison_readiness"
    if "faq structure" in text:
        return "semantic_depth"
    if "editorial linking" in text:
        return "trust_content"
    if "natural language search" in text or "filter integrity" in text or "zero results handling" in text:
        return "search_facets"
    if "search logging" in text:
        return "manual_search_analytics"
    if "clean html" in text:
        return "schema_alignment"
    if "product titles" in text or "product description" in text:
        return "product_metadata"
    if "schema" in text or "json" in text:
        return "schema_depth" if "variant schema" in text or "structured data" in text else "product_schema_presence"
    if "faq" in text or "reviews" in text or "guides" in text:
        return "semantic_depth"
    if "robots" in text or "sitemap" in text:
        return "robots"
    if "url" in text or "canonical" in text:
        return "url_structure"
    if "storefront api" in text:
        return "out_of_scope"
    if "external feeds" in text:
        return "machine_feeds"
    if "webhooks" in text:
        return "webhook_setup"
    if "data ownership" in text:
        return "manual_data_ownership"
    if "llms" in text:
        return "llms_txt"
    if "custom mcp" in text:
        return "custom_mcp"
    if "mcp" in text:
        return "mcp"
    if "regional context" in text:
        return "regional_signals"
    return "out_of_scope"


def classify_internal_question(question: str) -> str:
    q = (question or "").lower()
    if "master” version of each product" in q or "master version of each product" in q:
        return "url_structure"
    if "product titles clear and descriptive" in q:
        return "product_metadata"
    if "separate factual information" in q:
        return "semantic_depth"
    if "differences between variants" in q:
        return "schema_depth"
    if "collections represent how customers shop" in q:
        return "collection_content"
    if "related products” defined intentionally" in q or "related products" in q or "upsells and alternatives" in q:
        return "related_products"
    if "recommendations based on data" in q or "staff member give the same recommendation" in q:
        return "comparison_readiness"
    if "pros and cons between variants" in q:
        return "comparison_readiness"
    if "option values consistent" in q:
        return "option_consistency"
    if "fields/metafields" in q:
        return "metafield_quality"
    if "metafields being used" in q:
        return "metafield_quality"
    if "metafield namespaces" in q:
        return "metafield_namespaces"
    if "metaobjects used" in q:
        return "metaobject_usage"
    if "product categories" in q:
        return "taxonomy_assignment"
    if "collections represent how customers shop" in q:
        return "collection_content"
    if "why a product is in a collection" in q:
        return "collection_logic"
    if "not suitable for" in q:
        return "follow_up_readiness"
    if "legal, safety, or compliance restrictions" in q:
        return "follow_up_readiness"
    if "product media labelled clearly" in q:
        return "image_alt"
    if "draft, active, or archived" in q:
        return "manual_product_lifecycle"
    if "inventory tracked at variant level" in q:
        return "inventory_tracking"
    if "source of truth for stock" in q or "up-to-date is stock data" in q or "stock updates are instant or delayed" in q:
        return "manual_inventory_freshness"
    if "continue selling when out of stock" in q or "overselling" in q or "backorders" in q or "pre-orders" in q:
        return "inventory_policy"
    if "multiple locations or warehouses" in q:
        return "inventory_locations"
    if "fulfilment limits" in q:
        return "inventory_locations"
    if "stock messaging stay consistent" in q or "available right now" in q:
        return "manual_inventory_experience"
    if "webhooks reliable" in q:
        return "webhook_setup"
    if "data changes be traced back" in q or "system owns derived or calculated data" in q or "data changes logged and auditable" in q:
        return "manual_data_ownership"
    if "important business logic hidden in theme files" in q or "business logic hidden in theme" in q:
        return "theme_architecture"
    if "theme vs headless vs apps" in q or "online store 2.0 structure" in q:
        return "theme_architecture"
    if "filters map cleanly to real product attributes" in q:
        return "search_facets"
    if "why certain products appear in results" in q:
        return "manual_search_analytics"
    if "accurately reference the correct market or region" in q:
        return "regional_signals"
    if "duplicate content intentionally managed" in q:
        return "url_structure"
    if "indexation rules" in q:
        return "robots"
    if "what search engines see match what ai systems infer" in q or "discrepancies between human-visible content and machine-readable content" in q:
        return "schema_alignment"
    if "structured data (json-ld) accurately reflect" in q or "represented correctly in structured data" in q:
        return "schema_alignment"
    if any(term in q for term in ["fabric", "fit", "weather", "style", "material", "care instructions"]):
        return "clothing_attribute_depth"
    if any(term in q for term in ["mcp", "rpc", "manifest"]):
        return "mcp_depth"
    if any(term in q for term in ["intentional and documented", "prices, availability, and variants represented correctly"]):
        return "schema_depth"
    if any(term in q for term in ["robots", "sitemap", "canonical", "indexation", "semantic html", "machine-readable", "structured data", "json-ld"]):
        if "canonical" in q:
            return "url_structure"
        if "semantic html" in q or "machine-readable" in q:
            return "schema_alignment"
        if "structured data" in q or "json-ld" in q:
            return "schema_depth"
        return "robots" if "robots" in q else "sitemap"
    if any(term in q for term in ["product titles", "product descriptions", "main image", "detail image", "specification image"]):
        return "product_metadata" if "image" not in q else "image_alt"
    if any(term in q for term in ["metafield", "metaobject", "shopify product categories", "taxonomy", "variant", "master", "collections represent", "lifecycle", "draft, active, or archived"]):
        if "metafield namespaces" in q:
            return "metafield_namespaces"
        if "metaobject" in q:
            return "metaobject_usage"
        if "product categories" in q or "taxonomy" in q:
            return "taxonomy_assignment"
        if "lifecycle" in q or "draft, active, or archived" in q:
            return "manual_product_lifecycle"
        if "collections represent" in q:
            return "collection_logic"
        if "variant" in q:
            return "option_consistency"
        return "metafield_quality"
    if any(term in q for term in ["inventory", "stock", "backorders", "pre-orders", "checkout", "discount", "shipping", "payment", "tax", "webhooks", "admin api", "functions", "custom apps", "deployments", "headless", "theme vs", "owners and permissions", "roll back changes"]):
        if "inventory tracked" in q:
            return "inventory_tracking"
        if "backorders" in q or "pre-orders" in q or "continue selling" in q or "overselling" in q:
            return "inventory_policy"
        if "stock data" in q or "source of truth" in q or "instant or delayed" in q:
            return "manual_inventory_freshness"
        if "checkout" in q or "discount" in q or "shipping" in q or "payment" in q or "tax" in q:
            return "manual_commerce_logic"
        if "webhooks" in q:
            return "webhook_setup"
        if "headless" in q or "theme vs" in q or "custom apps" in q or "owners and permissions" in q or "roll back changes" in q or "deployments" in q:
            return "theme_architecture"
        return "manual_data_ownership"
    if any(term in q for term in ["search using full sentences", "no results", "filters map cleanly", "search terms and behaviours logged", "poor searches"]):
        return "manual_search_analytics" if "logged" in q or "poor searches" in q else "search_facets"
    if "how often content should be reviewed or updated" in q:
        return "manual_content_governance"
    if "seo changes checked against how ai agents behave" in q or "semantic/data changes reviewed for impact on recommendations" in q:
        return "manual_ai_governance"
    if "tools in place to see what ai agents are doing" in q:
        return "manual_ai_tooling"
    if any(term in q for term in ["faqs", "comparison content", "editorial content", "outdated content", "should i buy this"]):
        if "editorial content" in q:
            return "trust_content"
        if "outdated content" in q:
            return "manual_content_governance"
        if "comparison content" in q:
            return "comparison_readiness"
        return "semantic_depth"
    if any(term in q for term in [
        "traffic coming from ai agents",
        "agent succeeded",
        "ai-led journeys",
        "uncertain or incorrect answers",
        "problem areas clearly visible",
        "shared responsibility",
        "named owner",
        "discovery failures be traced",
        "performance reviewed",
        "example questions",
        "meaning drifts",
        "refreshing ai data",
        "common journeys tested again",
        "ai behaviour be rolled back",
        "tooling",
        "changes be tested safely",
        "costs of running ai",
        "tooling costs separate",
        "interpreting ai performance data",
        "measured before changes were made",
        "compared to the starting point",
        "success, confidence, and failure rates",
        "results tracked consistently",
        "internal benchmarks",
        "stakeholders",
    ]):
        if any(term in q for term in ["traffic coming from ai agents", "agent succeeded", "ai-led journeys", "uncertain or incorrect answers"]):
            return "manual_agent_analytics"
        if "problem areas clearly visible" in q:
            return "manual_agent_analytics"
        if any(term in q for term in ["shared responsibility", "named owner", "discovery failures be traced", "seo changes checked against how ai agents behave", "semantic/data changes reviewed for impact on recommendations"]):
            return "manual_ai_governance"
        if any(term in q for term in ["performance reviewed", "example questions", "meaning drifts", "refreshing ai data", "common journeys tested again", "ai behaviour be rolled back"]):
            return "manual_ai_quality"
        if any(term in q for term in ["tooling", "tools in place to see what ai agents are doing", "changes be tested safely", "costs of running ai", "tooling costs separate", "interpreting ai performance data"]):
            return "manual_ai_tooling"
        if any(term in q for term in ["measured before changes were made", "compared to the starting point", "success, confidence, and failure rates", "results tracked consistently", "internal benchmarks", "benchmarks", "stakeholders"]):
            return "manual_ai_benchmarks"
        return "manual_ai_governance"
    if any(term in q for term in ["brand appear consistently", "brand information", "clean, structured feeds", "shared directly with partners", "single trusted source", "usage rights", "regional context", "correct market or region", "quote facts about the brand"]):
        if "brand appear consistently" in q or "brand information" in q:
            return "manual_ai_ecosystem"
        if "region" in q or "regional" in q:
            return "regional_signals"
        if "feeds" in q or "shared directly" in q:
            return "machine_feeds"
        if "single trusted source" in q:
            return "trust_content"
        return "manual_ai_ecosystem"
    if any(term in q for term in ["important information visible in the page markup", "what search engines see", "discrepancies between human-visible content and machine-readable content"]):
        return "schema_alignment"
    if any(term in q for term in ["vector database"]):
        return "out_of_scope"
    if "products stored as “what they are”" in q or "products stored as \"what they are\"" in q or "products stored as 'what they are'" in q:
        return "manual_data_ownership"
    if any(term in q for term in ["compare two products side by side", "pros and cons between variants", "use cases", "best suited for", "limitations or caveats", "follow-up questions", "recommended over another", "recommendations based on data", "staff member give the same recommendation"]):
        if "recommended over another" in q or "recommendations based on data" in q or "staff member give the same recommendation" in q:
            return "comparison_readiness"
        return "comparison_readiness"
    if "shared directly with partners" in q or "clean, structured feeds" in q:
        return "machine_feeds"
    if "allowed to be used by ai tools" in q or "usage rights documented" in q:
        return "llms_txt"
    if "regional context" in q:
        return "regional_signals"
    return "out_of_scope"


def readiness_cell_text(site_result: Dict[str, Any], visibility_stats: Dict[str, Any] | None) -> str:
    if visibility_stats is not None:
        return f"{visibility_stats['mention_rate']:.0f}% mention | avg rank {visibility_stats['avg_rank'] or '-'} | score {visibility_stats['visibility_score']:.1f}"
    if site_result["status"] == "OUT OF SCOPE":
        return "OUT OF SCOPE / Not publicly visible"
    if site_result["status"] == "MANUAL":
        return "NO PUBLIC EVIDENCE / Separate review"
    return f"{site_result['status']} ({site_result['score']}) | {site_result['evidence'][:60]}"


def readiness_visibility_text(check_key: str, insight_map: Dict[str, Dict[str, Any]]) -> str:
    if check_key == "visibility_brand":
        item = insight_map["brand"]
        note = f" Remark: {item['note']}" if item.get("note") else ""
        return (
            f"{item['mention_rate']:.0f}% mention | avg rank {item['avg_rank'] or '-'} | "
            f"top 3 {item['top3_rate']:.0f}% | {item['remark']}{note}"
        )
    if check_key == "visibility_product":
        item = insight_map["product"]
        note = f" Remark: {item['note']}" if item.get("note") else ""
        return f"{item['product_citation_rate']:.0f}% PDP citation rate | {item['remark']}{note}"
    if check_key == "visibility_citation":
        item = insight_map["citation"]
        note = f" Remark: {item['note']}" if item.get("note") else ""
        return (
            f"{item['citation_rate']:.0f}% site citation | {item['product_citation_rate']:.0f}% PDP citation | "
            f"{item['remark']}{note}"
        )
    if check_key == "visibility_framing":
        item = insight_map["framing"]
        note = f" Remark: {item['note']}" if item.get("note") else ""
        return (
            f"{item['dominant_sentiment']} framing | +{item['positive']} / "
            f"~{item['neutral']} / -{item['negative']} | {item['remark']}{note}"
        )
    if check_key == "visibility_competitors":
        item = insight_map["competitors"]
        note = f" Remark: {item['note']}" if item.get("note") else ""
        competitor = item["top_competitor"] or "none"
        return (
            f"Top competitor: {competitor} | mention rate {item['top_competitor_mention_rate']:.0f}% | "
            f"{item['remark']}{note}"
        )
    return "No visibility insight"


def site_check_or_default(site_checks: Dict[str, Dict[str, Any]], key: str) -> Dict[str, Any]:
    return site_checks.get(key) or site_checks["out_of_scope"]


def out_of_scope_result(generator_module) -> Dict[str, Any]:
    return generator_module.result(
        "OUT OF SCOPE",
        0,
        "This check depends on platform, operational, analytics, or process evidence that is not visible in the public storefront render.",
        "Keep this outside public storefront scoring or review it in a separate internal validation pass.",
        observability="Public render only",
        sources="Out of scope",
    )


def fill_visibility_sheet(master_ws, report_ws) -> None:
    copy_sheet_exact(report_ws, master_ws)


def fill_overall_reporting(ws, visibility_rows: Dict[str, List[Dict[str, Any]]]) -> None:
    blocks = {
        "trailberg.com": 2,
        "lorenzo.world": 9,
        "dreamisfree.com": 16,
    }
    for domain, start_col in blocks.items():
        rows = visibility_rows[domain]
        target_brand = rows[0]["target_brand_name"] if rows else SITE_LABELS[domain]
        metrics = compute_visibility_metrics(rows, target_brand)
        seen_platforms = []
        for row in rows:
            platform = str(row.get("platform") or "").strip()
            if platform and platform not in seen_platforms and str(row.get("status") or "").strip().lower() != "error":
                seen_platforms.append(platform)
        platforms = seen_platforms[:3]
        if len(platforms) < 3:
            for fallback in ["All Platforms"]:
                if fallback not in platforms:
                    platforms.append(fallback)
                if len(platforms) == 3:
                    break
        ws.cell(2, start_col, f"https://www.{domain}/" if domain != "dreamisfree.com" else "https://dreamisfree.com/")
        row_ptr = 4
        for platform in platforms:
            brands = metrics.get(platform, {}).get("top_brands", [])
            padded = brands + [{"brand": None, "mention_rate": None, "avg_rank": None, "top3_rate": None, "visibility_score": None}] * (5 - len(brands))
            for item in padded[:5]:
                ws.cell(row_ptr, start_col, platform)
                ws.cell(row_ptr, start_col + 1, item["brand"])
                ws.cell(row_ptr, start_col + 2, item["mention_rate"])
                ws.cell(row_ptr, start_col + 3, item["avg_rank"])
                ws.cell(row_ptr, start_col + 4, item["top3_rate"])
                ws.cell(row_ptr, start_col + 5, item["visibility_score"])
                row_ptr += 1


def fill_readiness_sheet(ws, site_checks: Dict[str, Dict[str, Dict[str, Any]]], visibility_rows: Dict[str, List[Dict[str, Any]]]) -> None:
    for idx, domain in enumerate(SITE_ORDER, start=7):
        cell = ws.cell(1, idx, SITE_LABELS[domain])
        clone_style(ws["F1"], cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 22
    ws.column_dimensions["I"].width = 22

    visibility_summary = {}
    for domain, rows in visibility_rows.items():
        target_brand = rows[0]["target_brand_name"] if rows else SITE_LABELS[domain]
        target_domain = rows[0].get("target_brand_website") if rows else domain
        visibility_summary[domain] = summarize_visibility_insights(rows, target_brand, target_domain or domain)

    row = 2
    while row <= ws.max_row and any(ws.cell(row, col).value for col in range(2, 7)):
        domain_text = ws.cell(row, 3).value or ""
        what_check = ws.cell(row, 4).value or ""
        check_key = classify_readiness_row(domain_text, what_check)
        if check_key in PRIVATE_CHECK_KEYS:
            ws.delete_rows(row, 1)
            continue
        for col_offset, domain in enumerate(SITE_ORDER, start=7):
            if check_key.startswith("visibility_"):
                text = readiness_visibility_text(check_key, visibility_summary[domain])
            else:
                text = readiness_cell_text(site_checks[domain][check_key], None)
            cell = ws.cell(row, col_offset, text)
            clone_style(ws["F2"], cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1


def fill_internal_sheet(ws, site_checks: Dict[str, Dict[str, Dict[str, Any]]]) -> None:
    extra_headers = {19: "Trailberg", 20: "Lorenzo", 21: "Dream Is Free"}
    for col, value in extra_headers.items():
        cell = ws.cell(1, col, value)
        clone_style(ws["R1"], cell)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for col in ["S", "T", "U"]:
        ws.column_dimensions[col].width = 18

    row = 2
    while row <= ws.max_row and ws.cell(row, 5).value:
        question = str(ws.cell(row, 5).value)
        check_key = classify_internal_question(question)
        if check_key in PRIVATE_CHECK_KEYS:
            ws.delete_rows(row, 1)
            continue
        items = [site_check_or_default(site_checks[domain], check_key) for domain in SITE_ORDER]
        combined_score = average_score(items)
        ws.cell(row, 8, summary_text(items))
        ws.cell(row, 9, combined_score if combined_score else None)
        ws.cell(row, 10, " | ".join(f"{SITE_LABELS[d]}: {site_check_or_default(site_checks[d], check_key)['evidence']}" for d in SITE_ORDER))
        ws.cell(row, 11, target_outcome(check_key))
        ws.cell(row, 12, " | ".join(dict.fromkeys(item["recommendation"] for item in items)))
        ws.cell(row, 13, vector_db_flag(check_key))
        ws.cell(row, 14, ws.cell(row, 3).value)
        ws.cell(row, 15, recommended_effort(items))
        ws.cell(row, 16, business_impact(check_key))
        ws.cell(row, 17, target_outcome(check_key))
        if all(item["status"] == "OUT OF SCOPE" for item in items):
            status = "OUT OF SCOPE"
        elif all(item["status"] == "MANUAL" for item in items):
            status = "MANUAL"
        else:
            status = rating(combined_score)
        ws.cell(row, 18, status)
        for col, value in zip([19, 20, 21], [compact_site_summary(site_check_or_default(site_checks[d], check_key)) for d in SITE_ORDER]):
            ws.cell(row, col, value)
        for col in range(8, 22):
            cell = ws.cell(row, col)
            clone_style(ws.cell(2, min(col, 18)), cell)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row += 1


def merge_admin_checks(
    site_checks: Dict[str, Dict[str, Dict[str, Any]]],
    admin_payload: Dict[str, Any] | None,
) -> None:
    if not admin_payload:
        return
    for site in admin_payload.get("results", []):
        domain = site.get("domain")
        if domain not in site_checks:
            continue
        site_checks[domain].update(site.get("checks", {}))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--visibility-report", required=True)
    parser.add_argument("--audit-json", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH))
    parser.add_argument("--admin-audit-json")
    args = parser.parse_args()

    load_config(args.config)

    generator_module = load_existing_module("scripts/generate_shopify_audit_workbook.py")

    with open(args.audit_json, "r", encoding="utf-8") as fh:
        audit_payload = json.load(fh)
    visibility_wb = load_workbook(args.visibility_report)
    visibility_rows = load_visibility_rows(visibility_wb)

    site_checks: Dict[str, Dict[str, Dict[str, Any]]] = {}
    site_results_by_domain = {result["domain"]: result for result in audit_payload["results"]}
    for domain in SITE_ORDER:
        metrics = generator_module.build_metrics(site_results_by_domain[domain])
        site_checks[domain] = {
            check["key"]: generator_module.eval_check(check["key"], metrics)
            for check in generator_module.DETAIL_CHECKS
        }
        site_checks[domain].update(manual_review_checks(generator_module))
        site_checks[domain]["out_of_scope"] = out_of_scope_result(generator_module)

    admin_payload = None
    if args.admin_audit_json:
        with open(args.admin_audit_json, "r", encoding="utf-8") as fh:
            admin_payload = json.load(fh)
    merge_admin_checks(site_checks, admin_payload)

    wb = load_workbook(args.template)

    if "Summary" in wb.sheetnames:
        del wb["Summary"]
    wb.create_sheet("Summary", 0)
    generator_module.write_summary_sheet(
        wb["Summary"],
        [site_results_by_domain[domain] for domain in SITE_ORDER],
        [site_checks[domain] for domain in SITE_ORDER],
        audit_payload.get("generated_at", datetime.utcnow().isoformat()),
    )

    fill_overall_reporting(wb["Overall Reporting "], visibility_rows)
    fill_readiness_sheet(wb["Shopify_Agent Readiness"], site_checks, visibility_rows)
    fill_internal_sheet(wb["Shopify_Internal Structure Audi"], site_checks)

    for domain in SITE_ORDER:
        fill_visibility_sheet(wb[VISIBILITY_SHEET_MAP[domain]], visibility_wb[VISIBILITY_REPORT_MAP[domain]])

    wb.properties.modified = datetime.utcnow()
    wb.properties.creator = "Codex"
    wb.properties.lastModifiedBy = "Codex"

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    main()
