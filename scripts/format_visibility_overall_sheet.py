#!/usr/bin/env python3

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def copy_cell_style(src, dst) -> None:
    dst._style = copy(src._style)
    dst.number_format = src.number_format
    dst.alignment = copy(src.alignment)
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.protection = copy(src.protection)


def detect_site_blocks(values: dict[tuple[int, int], object], max_col: int) -> int:
    # Row 3 contains "Platform" at the start of each site block.
    row = 3
    starts = []
    for col in range(1, max_col + 1):
        value = str(values.get((row, col), "") or "").strip().lower()
        if value == "platform":
            starts.append(col)
    return len(starts) if starts else 0


def apply_overall_template_format(
    workbook_path: Path,
    template_path: Path,
    target_sheet_name: str = "Overall Reporting",
    template_sheet_name: str = "Overall Reporting ",
) -> None:
    wb = load_workbook(workbook_path)
    if target_sheet_name not in wb.sheetnames:
        raise ValueError(f"Target sheet not found: {target_sheet_name}")

    template_wb = load_workbook(template_path)
    if template_sheet_name not in template_wb.sheetnames:
        raise ValueError(f"Template sheet not found: {template_sheet_name}")

    target_ws = wb[target_sheet_name]
    template_ws = template_wb[template_sheet_name]

    values = {}
    max_row = target_ws.max_row
    max_col = target_ws.max_column
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            values[(row, col)] = target_ws.cell(row, col).value

    site_blocks = detect_site_blocks(values, max_col)
    if site_blocks <= 0:
        raise ValueError("Could not detect site blocks from current overall sheet values.")

    sheet_index = wb.sheetnames.index(target_sheet_name)
    del wb[target_sheet_name]
    styled_ws = wb.create_sheet(title=target_sheet_name, index=sheet_index)

    # Mirror sheet-level settings.
    styled_ws.freeze_panes = template_ws.freeze_panes
    styled_ws.sheet_view.showGridLines = template_ws.sheet_view.showGridLines

    # Copy row heights from template for visible range.
    template_max_row = template_ws.max_row
    final_max_row = max(max_row, template_max_row)
    for row_idx in range(1, final_max_row + 1):
        src_idx = row_idx if row_idx <= template_max_row else template_max_row
        src_dim = template_ws.row_dimensions[src_idx]
        dst_dim = styled_ws.row_dimensions[row_idx]
        dst_dim.height = src_dim.height
        dst_dim.hidden = src_dim.hidden

    # Column 1 (left gutter) style/width from template col A.
    src_col_a = 1
    styled_ws.column_dimensions["A"].width = template_ws.column_dimensions["A"].width
    for row_idx in range(1, final_max_row + 1):
        copy_cell_style(template_ws.cell(min(row_idx, template_max_row), src_col_a), styled_ws.cell(row_idx, 1))

    # Repeat block formatting from template B:H for each detected site block.
    # Block width = 7 columns: B..G content + H spacer.
    for block_idx in range(site_blocks):
        target_start_col = 2 + (block_idx * 7)
        for offset in range(7):
            src_col = 2 + offset
            dst_col = target_start_col + offset
            src_letter = get_column_letter(src_col)
            dst_letter = get_column_letter(dst_col)
            styled_ws.column_dimensions[dst_letter].width = template_ws.column_dimensions[src_letter].width

            for row_idx in range(1, final_max_row + 1):
                src_row = row_idx if row_idx <= template_max_row else template_max_row
                copy_cell_style(template_ws.cell(src_row, src_col), styled_ws.cell(row_idx, dst_col))

        # Merge row-2 website header area across the 6 data columns.
        styled_ws.merge_cells(
            start_row=2,
            start_column=target_start_col,
            end_row=2,
            end_column=target_start_col + 5,
        )

    # Re-apply values from existing consolidated sheet.
    for (row, col), value in values.items():
        if value is None:
            continue
        styled_ws.cell(row, col, value)

    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply template formatting to consolidated visibility Overall Reporting sheet.")
    parser.add_argument("--workbook", required=True, help="Merged visibility workbook path")
    parser.add_argument("--template", required=True, help="Master template workbook path")
    parser.add_argument("--target-sheet", default="Overall Reporting", help="Target sheet name in merged workbook")
    parser.add_argument("--template-sheet", default="Overall Reporting ", help="Source sheet name in template workbook")
    args = parser.parse_args()

    apply_overall_template_format(
        workbook_path=Path(args.workbook),
        template_path=Path(args.template),
        target_sheet_name=args.target_sheet,
        template_sheet_name=args.template_sheet,
    )
    print(f"Applied template formatting to '{args.target_sheet}' in {args.workbook}")


if __name__ == "__main__":
    main()
